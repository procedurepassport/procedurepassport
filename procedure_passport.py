import streamlit as st
import time
import pandas as pd
import uuid
import datetime
import io
import json
import html
import re
import hashlib
import secrets
import gspread
from gspread_dataframe import get_as_dataframe, set_with_dataframe
from google.oauth2.service_account import Credentials
import numpy as np


st.set_page_config(
    page_title="Procedure Passport",
    page_icon="🩺",
    layout="wide",
)

# ─────────────────────────────────────────────
# QUERY PARAMS  (magic link routing)
# ─────────────────────────────────────────────
query_params = st.query_params

# Only auto-route on the first load; once submitted we stay on the confirmation page.
if (
    query_params.get("mode") == "attending"
    and st.session_state.get("page", "login") not in ("attending_confirmation",)
    and not st.session_state.get("_magic_routed")
):
    st.session_state["page"]              = "attending_assessment"
    st.session_state["resident"]          = query_params.get("resident", "")
    st.session_state["procedure_id"]      = query_params.get("procedure_id", "")
    st.session_state["specialty_id"]      = query_params.get("specialty_id", "")
    st.session_state["attending_name"]    = query_params.get("attending_name", "")
    st.session_state["draft_id"]          = query_params.get("draft_id", "")
    # Only carried by the blank-link flow (the self-assess/pre-filled flow
    # gets its date from the draft instead) — the date the resident chose
    # on the Start page before generating the blank link, attached to the
    # submission silently; there's no UI on the attending page to view or
    # edit it.
    st.session_state["attending_link_date"] = query_params.get("date", "")
    st.session_state["_magic_routed"]     = True

# ─────────────────────────────────────────────
# SESSION STATE DEFAULTS
# ─────────────────────────────────────────────
_defaults: dict = {
    "page":                    "login",
    "resident":                None,
    "resident_name":           "",
    "scores":                  {},
    "date":                    datetime.date.today(),
    "notes":                   "",
    "improve":                 "",
    "how":                     "",
    "current_case_id":         None,
    "attending_submission":    None,   # filled after magic-link submit
    "generated_magic_link":    None,   # filled after Generate Magic Link
    "draft_id":                "",
    "assessment_mode":         "together",  # "together" or "self", set from Start
    "blank_magic_link":        None,   # filled after Generate a Blank Magic Link
    "attending_link_date":     "",     # resident's chosen date, carried by a blank magic link
    "last_assessment_type":    None,   # "Assessed Together" or "Self-Assessment"
    "role":                    None,   # "resident", "admin", or "attending" — set at login
    "attending_login_email":   "",     # set only when role == "attending" (their own account)
    "attending_login_name":    "",
    "attending_login_id":      "",
    "attending_login_specialty_id": "",
}
for _k, _v in _defaults.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
ADMINS = ["procedurepassport@gmail.com"]

RATING_OPTIONS = ["Not Assessed", "Shown/Told", "Not Yet", "Steer", "Prompt", "Back up", "Auto"]
RATING_TO_NUM  = {
    "Not Assessed": -1,
    "Shown/Told":    0,
    "Not Yet":       1,
    "Steer":         2,
    "Prompt":        3,
    "Back up":       4,
    "Auto":          5,
}
RATING_HEX = {
    "Not Assessed": "#FAFAFA",  # near-white — explicitly rated as not assessed
    "Shown/Told":   "#9E9E9E",  # dark gray — explicitly shown or told
    "Not Yet":      "#5B8DB8",
    "Steer":        "#FF944D",
    "Prompt":       "#FFD633",
    "Back up":      "#99E699",
    "Auto":         "#33CC33",
}
RATING_COLOR = {
    k: f"background-color:{v}; color:{'white' if k in ('Not Yet','Auto') else 'black'};"
    for k, v in RATING_HEX.items()
}

def fmt_date(d):
    """Format a date value as MM-DD-YYYY; pass through non-date strings unchanged."""
    try:
        if pd.isna(d):
            return ""
    except TypeError:
        pass
    try:
        return pd.Timestamp(d).strftime("%m-%d-%Y")
    except Exception:
        return str(d)


def _norm_id(series: pd.Series) -> pd.Series:
    """Normalise a case_id Series to clean strings regardless of pandas version.

    pandas 3.x can infer all-digit hex IDs as float64, making astype(str)
    produce "123456789012.0" while the other sheet retains "123456789012".
    The three-step chain below is safe for every dtype:
      float64  123456789012.0  → "123456789012.0" → strip → remove .0 → "123456789012"
      int64    123456789012    → "123456789012"   → strip → no-op      → "123456789012"
      object   "abc123def456"  → "abc123def456"   → strip → no-op      → "abc123def456"
    """
    return (series.astype(str)
                  .str.strip()
                  .str.replace(r"\.0$", "", regex=True))


COMPLEXITY_HEX = {
    "Straight Forward": "#C8E6C9",
    "Moderate":         "#FFF59D",
    "Complex":          "#FFAB91",
}
O_SCORE_HEX = {
    "1": "#378ADD",
    "2": "#FF944D",
    "3": "#FFD633",
    "4": "#99E699",
    "5": "#33CC33",
}
O_SCORE_OPTIONS = [
    "— Make a selection —",
    "1 - Not Yet",
    "2 - Steer",
    "3 - Prompt",
    "4 - Backup",
    "5 - Auto",
]

SHEET_RESIDENTS  = "residents"
SHEET_ATTENDINGS = "attendings"
SHEET_PROCEDURES = "procedures"
SHEET_STEPS      = "steps"
SHEET_CASES      = "cases"
SHEET_SCORES     = "scores"
SHEET_SPECIALTY  = "specialties"
SHEET_DRAFTS     = "drafts"
SHEET_AUTH       = "auth"

# Password auth, one row per email (resident or admin) that has ever set a
# password. Deliberately its own sheet, not columns on `residents` — the
# admin account isn't a residents-sheet row at all, and keeping credential
# material out of the general roster is good hygiene regardless.
AUTH_COLS = ["email", "password_hash", "password_salt", "created_at"]
PBKDF2_ITERATIONS = 200_000

# Pre-filled magic-link drafts: a resident's in-progress assessment, saved
# so the attending's link can carry a short draft_id instead of embedding
# every field's value in the URL itself.
DRAFT_COLS = [
    "draft_id", "resident_email", "date", "specialty_id", "procedure_id",
    "attending_id", "case_complexity", "case_preparation",
    "overall_performance", "robo_type", "improve", "how", "notes",
    "scores_json", "created_at",
]

# ─────────────────────────────────────────────
# GOOGLE SHEETS HELPERS
# ─────────────────────────────────────────────

@st.cache_resource(show_spinner=False)
def get_gs_client():
    """Authorized gspread client — cached for the entire app session."""
    creds = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ],
    )
    return gspread.authorize(creds)


def get_sheet(sheet_name: str):
    """Return a gspread worksheet, creating it if missing."""
    try:
        gc = get_gs_client()
        sh = gc.open_by_key(st.secrets["GOOGLE_SHEET_KEY"])
        try:
            return sh.worksheet(sheet_name)
        except gspread.exceptions.WorksheetNotFound:
            return sh.add_worksheet(title=sheet_name, rows="500", cols="26")
    except Exception as exc:
        raise ConnectionError(f"Cannot reach Google Sheets: {exc}") from exc


@st.cache_data(ttl=300, show_spinner=False)
def read_sheet_df(sheet_name: str, expected_cols=None) -> pd.DataFrame:
    """Cached worksheet read (300 s TTL).  Returns empty DF if sheet is blank."""
    ws  = get_sheet(sheet_name)
    df  = get_as_dataframe(ws, evaluate_formulas=True, header=0)
    df  = df.dropna(how="all")
    if df.empty and expected_cols:
        return pd.DataFrame(columns=expected_cols)
    if expected_cols:
        for col in expected_cols:
            if col not in df.columns:
                df[col] = pd.NA
        df = df[expected_cols]
    return df


def write_sheet_df(sheet_name: str, df: pd.DataFrame) -> None:
    """Overwrite a worksheet then clear all cached reads so the UI stays fresh."""
    ws = get_sheet(sheet_name)
    ws.clear()
    set_with_dataframe(ws, df, include_index=False, include_column_header=True)
    st.cache_data.clear()  # invalidate all read caches after every write


@st.cache_data(ttl=300, show_spinner=False)
def load_refs():
    """Load all reference tables in one shot (cached 300 s)."""
    def _safe(name, cols):
        try:
            return read_sheet_df(name, expected_cols=cols)
        except Exception:
            return pd.DataFrame(columns=cols)

    spec_df  = _safe(SHEET_SPECIALTY,  ["specialty_id",  "specialty_name"])
    proc_df  = _safe(SHEET_PROCEDURES, ["procedure_id",  "procedure_name", "specialty_id"])
    steps_df = _safe(SHEET_STEPS,      ["step_id",       "procedure_id",   "step_order", "step_name"])
    try:
        atnd_df = _read_attendings_df()
    except Exception:
        atnd_df = pd.DataFrame(columns=ATTENDING_COLS)
    return spec_df, proc_df, steps_df, atnd_df


# ─────────────────────────────────────────────
# DATA MUTATION HELPERS
# ─────────────────────────────────────────────

def _hash_password(password: str, salt_hex: str) -> str:
    """PBKDF2-HMAC-SHA256, hex-encoded. salt_hex is a hex string (not raw
    bytes) so it round-trips through Google Sheets as plain text."""
    return hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), bytes.fromhex(salt_hex), PBKDF2_ITERATIONS
    ).hex()


def get_password_row(email: str):
    """Return the auth row dict for this email, or None if no password has
    ever been set for it."""
    auth_df = read_sheet_df(SHEET_AUTH, expected_cols=AUTH_COLS)
    if auth_df.empty:
        return None
    email_norm = email.strip().lower()
    match = auth_df[auth_df["email"].astype(str).str.strip().str.lower() == email_norm]
    if match.empty:
        return None
    row = match.iloc[0]
    if pd.isna(row.get("password_hash")) or not str(row.get("password_hash", "")).strip():
        return None
    return {"password_hash": str(row["password_hash"]), "password_salt": str(row["password_salt"])}


def set_password(email: str, password: str) -> None:
    """Create or overwrite the stored password for this email."""
    auth_df = read_sheet_df(SHEET_AUTH, expected_cols=AUTH_COLS)
    email_norm = email.strip().lower()
    salt_hex = secrets.token_bytes(16).hex()
    password_hash = _hash_password(password, salt_hex)
    auth_df = auth_df[auth_df["email"].astype(str).str.strip().str.lower() != email_norm]
    auth_df = pd.concat([auth_df, pd.DataFrame([{
        "email":         email.strip(),
        "password_hash": password_hash,
        "password_salt": salt_hex,
        "created_at":    datetime.datetime.utcnow().isoformat(),
    }])], ignore_index=True)
    write_sheet_df(SHEET_AUTH, auth_df)


def verify_password(email: str, password: str) -> bool:
    row = get_password_row(email)
    if row is None:
        return False
    candidate = _hash_password(password, row["password_salt"])
    return secrets.compare_digest(candidate, row["password_hash"])


def clear_password(email: str) -> None:
    """Remove the stored password for this email — their next login will
    prompt them to set a new one. Silently no-ops if none is on file."""
    auth_df = read_sheet_df(SHEET_AUTH, expected_cols=AUTH_COLS)
    if auth_df.empty:
        return
    email_norm = email.strip().lower()
    remaining = auth_df[auth_df["email"].astype(str).str.strip().str.lower() != email_norm]
    if len(remaining) != len(auth_df):
        write_sheet_df(SHEET_AUTH, remaining)


def ensure_resident(email: str, name: str = "", specialty_id=None) -> None:
    cols = ["email", "name", "specialty_id", "created_at"]
    df   = read_sheet_df(SHEET_RESIDENTS, expected_cols=cols)
    if email not in df["email"].values:
        df = pd.concat([df, pd.DataFrame([{
            "email":        email,
            "name":         name,
            "specialty_id": specialty_id,
            "created_at":   datetime.datetime.utcnow().isoformat(),
        }])], ignore_index=True)
        write_sheet_df(SHEET_RESIDENTS, df)   # also clears cache


def ensure_attending(name: str, specialty_id: str, email: str = "") -> None:
    df = _read_attendings_df()
    if name not in df["attending_name"].values:
        att_id = "A_" + specialty_id + "_" + name.replace(" ", "_").upper()
        df = pd.concat([df, pd.DataFrame([{
            "attending_id":   att_id,
            "attending_name": name,
            "specialty_id":   specialty_id,
            "email":          email,
        }])], ignore_index=True)
        write_sheet_df(SHEET_ATTENDINGS, df)


def ensure_procedure(proc_id: str, proc_name: str, specialty_id: str, steps_list: list) -> None:
    proc_cols = ["procedure_id", "procedure_name", "specialty_id"]
    procs_df  = read_sheet_df(SHEET_PROCEDURES, expected_cols=proc_cols)
    if proc_id not in procs_df["procedure_id"].values:
        procs_df = pd.concat([procs_df, pd.DataFrame([{
            "procedure_id":   proc_id,
            "procedure_name": proc_name,
            "specialty_id":   specialty_id,
        }])], ignore_index=True)
        write_sheet_df(SHEET_PROCEDURES, procs_df)

    step_cols = ["step_id", "procedure_id", "step_order", "step_name"]
    steps_df  = read_sheet_df(SHEET_STEPS, expected_cols=step_cols)
    if not (steps_df["procedure_id"] == proc_id).any():
        new_steps = pd.DataFrame([{
            "step_id":      f"S_{proc_id}_{i+1:02d}",
            "procedure_id": proc_id,
            "step_order":   i + 1,
            "step_name":    step,
        } for i, step in enumerate(steps_list)])
        steps_df = pd.concat([steps_df, new_steps], ignore_index=True)
        write_sheet_df(SHEET_STEPS, steps_df)


def save_case(
    resident_email: str,
    date,
    specialty_id: str,
    procedure_id: str,
    attending_id: str,
    scores_dict: dict,
    notes: str = "",
    case_complexity=None,
    case_preparation=None,
    overall_performance=None,
    robo_type=None,
    improve: str = "",
    how: str = "",
    assessment_type: str = "",
) -> str:
    """Persist a case + its step scores; returns the new case_id.

    assessment_type distinguishes who the case record represents:
    "Self-Assessment" for a resident's own entry (Finish & Save, or the
    dual-save that happens when generating a pre-filled magic link) vs
    "Attending Evaluation" for the attending's magic-link submission.

    robo_type ("Xi"/"SP"/"DV5") is only meaningful for a robotic
    procedure (see _is_robotic_procedure) — None otherwise.
    """
    case_id   = uuid.uuid4().hex[:12]

    case_cols = ["case_id", "resident_email", "date", "specialty_id",
                 "procedure_id", "attending_id", "notes",
                 "case_complexity", "case_preparation", "overall_performance",
                 "robo_type", "improve", "how", "assessment_type"]
    cases_df  = read_sheet_df(SHEET_CASES, expected_cols=case_cols)
    cases_df  = pd.concat([cases_df, pd.DataFrame([{
        "case_id":             case_id,
        "resident_email":      resident_email,
        "date":                str(date),
        "specialty_id":        specialty_id,
        "procedure_id":        procedure_id,
        "attending_id":        attending_id,
        "notes":               notes,
        "case_complexity":     case_complexity,
        "case_preparation":    case_preparation,
        "overall_performance": overall_performance,
        "robo_type":           robo_type,
        "improve":             improve,
        "how":                 how,
        "assessment_type":     assessment_type,
    }])], ignore_index=True)
    write_sheet_df(SHEET_CASES, cases_df)  # clears cache

    score_cols = ["case_id", "step_id", "rating", "rating_num",
                  "case_complexity", "case_preparation", "overall_performance"]
    scores_df  = read_sheet_df(SHEET_SCORES, expected_cols=score_cols)
    # Normalise existing case_ids before concat so the written sheet is consistent.
    if not scores_df.empty:
        scores_df["case_id"] = _norm_id(scores_df["case_id"])
    new_rows   = [{
        "case_id":             case_id,
        "step_id":             step_id,
        "rating":              rating,
        "rating_num":          RATING_TO_NUM.get(rating),
        "case_complexity":     case_complexity,
        "case_preparation":    case_preparation,
        "overall_performance": overall_performance,
    } for step_id, rating in scores_dict.items()]
    scores_df  = pd.concat([scores_df, pd.DataFrame(new_rows)], ignore_index=True)
    write_sheet_df(SHEET_SCORES, scores_df)  # clears cache

    return case_id


def save_draft(
    resident_email: str,
    date,
    specialty_id: str,
    procedure_id: str,
    attending_id: str,
    scores_dict: dict,
    notes: str = "",
    case_complexity=None,
    case_preparation=None,
    overall_performance=None,
    robo_type=None,
    improve: str = "",
    how: str = "",
) -> str:
    """Save a resident's in-progress assessment as a pre-fill draft for a
    magic link; returns the draft_id to embed in the link's query string."""
    draft_id  = uuid.uuid4().hex[:12]
    drafts_df = read_sheet_df(SHEET_DRAFTS, expected_cols=DRAFT_COLS)
    drafts_df = pd.concat([drafts_df, pd.DataFrame([{
        "draft_id":             draft_id,
        "resident_email":       resident_email,
        "date":                 str(date),
        "specialty_id":         specialty_id,
        "procedure_id":         procedure_id,
        "attending_id":         attending_id,
        "case_complexity":      case_complexity,
        "case_preparation":     case_preparation,
        "overall_performance":  overall_performance,
        "robo_type":            robo_type,
        "improve":              improve,
        "how":                  how,
        "notes":                notes,
        "scores_json":          json.dumps(scores_dict),
        "created_at":           datetime.datetime.utcnow().isoformat(),
    }])], ignore_index=True)
    write_sheet_df(SHEET_DRAFTS, drafts_df)
    return draft_id


def load_draft(draft_id: str):
    """Fetch a pre-fill draft by id. Returns None if missing, blank, or the
    sheet can't be reached — callers should fall back to a blank form."""
    if not draft_id:
        return None
    try:
        drafts_df = read_sheet_df(SHEET_DRAFTS, expected_cols=DRAFT_COLS)
    except ConnectionError:
        return None
    if drafts_df.empty:
        return None
    drafts_df   = drafts_df.copy()
    drafts_df["draft_id"] = _norm_id(drafts_df["draft_id"])
    target      = _norm_id(pd.Series([draft_id])).iloc[0]
    match       = drafts_df[drafts_df["draft_id"] == target]
    if match.empty:
        return None
    row = match.iloc[0]

    def _clean(v):
        # Blank sheet cells round-trip through pandas as NaN (a float), not
        # "" — and NaN is truthy in Python, so a plain `v or ""` doesn't
        # catch it, leaving the literal text "nan" in text inputs/areas.
        return "" if pd.isna(v) else str(v)

    try:
        scores = json.loads(_clean(row.get("scores_json")) or "{}")
    except (TypeError, ValueError, json.JSONDecodeError):
        scores = {}
    return {
        "date":                 row.get("date"),
        "case_complexity":      row.get("case_complexity"),
        "case_preparation":     row.get("case_preparation"),
        "overall_performance":  row.get("overall_performance"),
        "robo_type":            row.get("robo_type"),
        "improve":              _clean(row.get("improve")),
        "how":                  _clean(row.get("how")),
        "notes":                _clean(row.get("notes")),
        "scores":               scores,
    }


def delete_draft(draft_id: str) -> None:
    """Remove a consumed draft. Cleanup only — never raises."""
    if not draft_id:
        return
    try:
        drafts_df = read_sheet_df(SHEET_DRAFTS, expected_cols=DRAFT_COLS)
        if drafts_df.empty:
            return
        drafts_df = drafts_df.copy()
        drafts_df["draft_id"] = _norm_id(drafts_df["draft_id"])
        target    = _norm_id(pd.Series([draft_id])).iloc[0]
        remaining = drafts_df[drafts_df["draft_id"] != target]
        if len(remaining) != len(drafts_df):
            write_sheet_df(SHEET_DRAFTS, remaining)
    except Exception:
        pass


# ─────────────────────────────────────────────
# STYLING HELPERS
# ─────────────────────────────────────────────

def style_df(df: pd.DataFrame, col: str):
    return df.style.map(lambda v: RATING_COLOR.get(v, ""), subset=[col])


def attending_display_name(attending_id: str, atnds_lookup: dict) -> str:
    """Resolve a display name from an attending_id, including magic_ IDs."""
    if attending_id in atnds_lookup:
        return atnds_lookup[attending_id]
    if isinstance(attending_id, str) and attending_id.startswith("magic_"):
        return attending_id[len("magic_"):].replace("_", " ")
    return attending_id or "Unknown"


ATTENDING_COLS = ["attending_id", "attending_name", "specialty_id", "email"]


def _read_attendings_df() -> pd.DataFrame:
    """Read the attendings sheet as ATTENDING_COLS, tolerating a sheet whose
    login-email column is still named "attending_email" (an older schema —
    some existing sheets/export snapshots use that name) instead of "email".
    Always returns a single, populated "email" column regardless of which
    name the live sheet actually has, preferring "email" when both are
    present. Every read of the attendings sheet should go through this,
    not a raw read_sheet_df(SHEET_ATTENDINGS, ...) call, so that whichever
    column the data is really in, it's picked up — and so any write that
    follows (which only ever writes ATTENDING_COLS) carries the value
    forward under "email" instead of silently dropping it."""
    df = read_sheet_df(SHEET_ATTENDINGS, expected_cols=ATTENDING_COLS + ["attending_email"])
    _email  = df["email"].fillna("").astype(str).str.strip()
    _legacy = df["attending_email"].fillna("").astype(str).str.strip()
    df["email"] = _email.where(_email != "", _legacy)
    return df[ATTENDING_COLS]


# Pinned procedures always come first, in this order (when present for
# the current specialty); everything else follows alphabetically. Matched
# case/whitespace-insensitively (normalizing runs of whitespace, including
# non-breaking spaces, to a single plain space) since a sheet value that
# differs from these only by case or stray spacing should still be
# recognized as the same procedure and pinned correctly.
_PINNED_PROCS = [
    "Robotic Surgical Skills Feedback",
    "Robotic Bedsiding",
    "Open Surgical Skills Feedback",
    "Endoscopic Surgical Skills Feedback",
]


def _norm_proc(name: str) -> str:
    return " ".join(name.replace("\xa0", " ").split()).casefold()


def _ordered_procedure_names(proc_map: dict) -> list:
    """Procedure names for a Procedure dropdown: the pinned procedures
    first (in _PINNED_PROCS order), then everything else alphabetically."""
    _pinned_rank = {_norm_proc(name): i for i, name in enumerate(_PINNED_PROCS)}
    return sorted(
        proc_map.keys(),
        key=lambda n: (
            _pinned_rank.get(_norm_proc(n), len(_PINNED_PROCS)),
            n if _norm_proc(n) not in _pinned_rank else "",
        ),
    )


def _is_robotic_procedure(procedure_name: str) -> bool:
    """True if a procedure's name suggests it's done on a robotic
    platform, so the assessment form should show the Xi/SP/DV5 robot
    picker (see render_robo_type_picker). "robotic"/"robo" match
    case-insensitively anywhere in the name; "RAL" is matched only as
    an exact-case substring — case-insensitively it would false-positive
    on ordinary words that happen to end in "ral", e.g. "General"."""
    name = str(procedure_name or "")
    lname = name.lower()
    return "robotic" in lname or "robo" in lname or "RAL" in name


def _on_robo_checkbox_change(value_key: str, widget_keys: dict, clicked_label: str) -> None:
    """Keeps the Xi/SP/DV5 checkbox trio behaving like a single-select
    group even though st.checkbox has no native radio-group mode:
    checking one unchecks the other two and becomes the recorded
    selection; trying to uncheck the only checked one is refused (there
    must always be exactly one) by immediately re-checking it."""
    if st.session_state[widget_keys[clicked_label]]:
        for label, k in widget_keys.items():
            if label != clicked_label:
                st.session_state[k] = False
        st.session_state[value_key] = clicked_label
    else:
        st.session_state[widget_keys[clicked_label]] = True


def render_robo_type_picker(value_key: str, default: str = "Xi") -> str:
    """Renders "Robot:" and the Xi/SP/DV5 checkboxes inline in one tight
    row (the .st-key-assess_robo_row CSS rule shrinks each column to its
    own content instead of stretching evenly across the full row, which
    is what st.columns() does by default) and returns the current
    selection (also left in st.session_state[value_key] for
    save_case()/save_draft() to read). `default` only takes effect the
    first time `value_key` is ever set for this session — e.g. seeded
    from a pre-fill draft's own "robo_type" on the attending's page —
    and is ignored on every later rerun once the form has actually been
    interacted with."""
    _labels = ["Xi", "SP", "DV5"]
    if st.session_state.get(value_key) not in _labels:
        st.session_state[value_key] = default if default in _labels else "Xi"
    _current = st.session_state[value_key]
    _widget_keys = {label: f"{value_key}_cb_{label}" for label in _labels}
    for label, k in _widget_keys.items():
        if k not in st.session_state:
            st.session_state[k] = (label == _current)
    with st.container(key="assess_robo_row"):
        _label_col, *_cb_cols = st.columns(1 + len(_labels))
        with _label_col:
            st.markdown("**Robot:**")
        for _col, label in zip(_cb_cols, _labels):
            with _col:
                st.checkbox(
                    label, key=_widget_keys[label],
                    on_change=_on_robo_checkbox_change,
                    args=(value_key, _widget_keys, label),
                )
    return st.session_state[value_key]


def show_gs_error(exc: Exception) -> None:
    st.error(
        "⚠️ **Could not reach Google Sheets.** "
        "Check your network connection or try refreshing the page.\n\n"
        f"_Details: {exc}_"
    )


# ─────────────────────────────────────────────
# NAV HELPER
# ─────────────────────────────────────────────
def go_to(page: str) -> None:
    st.session_state["page"] = page
    st.rerun()


# ─────────────────────────────────────────────
# RESIDENT DATA HELPERS
# (shared by a resident's own Comments/Cumulative dashboards and by an
# attending's Resident Dashboard, which shows the same views for a resident
# of the attending's choosing)
# ─────────────────────────────────────────────
def _build_comments(row) -> str:
    """Plain-text Comments value (used for the Excel export): the
    "In order to improve this: ..." sentence (if either field was
    answered), followed by the free-text notes."""
    imp = row["improve"].strip()
    how = row["how"].strip()
    parts = []
    if imp or how:
        parts.append(f"In order to improve this: {imp or '(blank)'}.\nDo this: {how or '(blank)'}.")
    if row["notes"].strip():
        parts.append(row["notes"].strip())
    return "\n\n".join(parts)


def _build_comments_html(row) -> str:
    """HTML Comments value for the on-screen table: "In order to
    improve this:" and "Do this:" are bold labels, each starting its
    own line; the resident's own answers are underlined, not bold."""
    imp = row["improve"].strip()
    how = row["how"].strip()
    parts = []
    if imp or how:
        imp_html = f"<u>{html.escape(imp)}</u>" if imp else "(blank)"
        how_html = f"<u>{html.escape(how)}</u>" if how else "(blank)"
        parts.append(f"<b>In order to improve this:</b> {imp_html}.<br><b>Do this:</b> {how_html}.")
    if row["notes"].strip():
        parts.append(html.escape(row["notes"].strip()).replace(chr(10), "<br>"))
    return "<br><br>".join(parts)


def _build_resident_comments_df(resident_email: str) -> pd.DataFrame:
    """Attending-confirmed cases with a comment (improve/how/notes) for one
    resident, as a Date/Procedure/Attending/Comments/Comments_html table
    sorted newest first. Self-assessments are excluded — same as the
    heatmap, these are the resident's own unverified entry, not an
    attending-confirmed one. Empty DataFrame (with those columns) if there's
    nothing to show."""
    _cols = ["Date", "Procedure", "Attending", "Comments", "Comments_html"]
    cases_df = read_sheet_df(
        SHEET_CASES,
        expected_cols=["case_id", "resident_email", "date", "specialty_id",
                       "procedure_id", "attending_id", "notes",
                       "case_complexity", "overall_performance", "assessment_type",
                       "improve", "how"],
    )
    procs_df = read_sheet_df(SHEET_PROCEDURES, expected_cols=["procedure_id", "procedure_name", "specialty_id"])
    atnds_df = _read_attendings_df()

    cases_df["case_id"] = _norm_id(cases_df["case_id"])
    cases_df = cases_df.drop_duplicates(subset=["case_id"])

    res_cases = cases_df[cases_df["resident_email"] == resident_email].copy()
    res_cases = res_cases[res_cases["assessment_type"].fillna("").astype(str).str.strip() != "Self-Assessment"]
    res_cases["notes"]   = res_cases["notes"].fillna("").astype(str)
    res_cases["improve"] = res_cases["improve"].fillna("").astype(str)
    res_cases["how"]     = res_cases["how"].fillna("").astype(str)
    res_cases = res_cases[
        (res_cases["notes"].str.strip() != "")
        | (res_cases["improve"].str.strip() != "")
        | (res_cases["how"].str.strip() != "")
    ]
    if res_cases.empty:
        return pd.DataFrame(columns=_cols)

    res_cases["comments_html"] = res_cases.apply(_build_comments_html, axis=1)
    res_cases["notes"] = res_cases.apply(_build_comments, axis=1)

    atnds_lookup = dict(zip(atnds_df["attending_id"], atnds_df["attending_name"]))
    res_cases["attending_name"] = res_cases["attending_id"].apply(
        lambda aid: attending_display_name(str(aid), atnds_lookup)
    )

    procs_dedup = procs_df.drop_duplicates(subset=["procedure_id"])
    merged = res_cases.merge(procs_dedup[["procedure_id", "procedure_name"]], on="procedure_id", how="left")
    merged = merged.rename(columns={
        "date":           "Date",
        "procedure_name": "Procedure",
        "attending_name": "Attending",
        "notes":          "Comments",
        "comments_html":  "Comments_html",
    })
    merged["_date_sort"] = pd.to_datetime(merged["Date"], errors="coerce")
    merged = merged[_cols + ["_date_sort"]].sort_values("_date_sort", ascending=False).drop(columns=["_date_sort"])
    merged["Date"] = merged["Date"].apply(fmt_date)
    return merged


def _render_comments_html_table(merged: pd.DataFrame, show_proc: bool, show_att: bool) -> None:
    """Render a Date/[Procedure]/[Attending]/Comments table as wrapped HTML
    (so the Comments column can wrap), with the same shrink-to-fit script
    the Comments Dashboard uses."""
    st.markdown("""
<style>
.comments-tbl {width:100%;border-collapse:collapse;font-size:0.88rem;}
.comments-tbl th {background:var(--secondary-background-color);padding:8px 10px;
    text-align:left;border-bottom:2px solid #ccc;font-weight:600;}
.comments-tbl td {padding:8px 10px;vertical-align:top;border-bottom:1px solid var(--secondary-background-color);}
.comments-tbl td.date-col, .comments-tbl td.attending-col, .comments-tbl td.procedure-col {
    white-space:nowrap;font-size:var(--cmts-sync-font, inherit);
}
.comments-tbl td.comments-col {white-space:pre-wrap;word-break:break-word;min-width:260px;}
</style>""", unsafe_allow_html=True)

    _rows_html = ""
    for _, r in merged.reset_index(drop=True).iterrows():
        _rows_html += (
            f"<tr>"
            f"<td class='date-col'>{html.escape(str(r['Date']))}</td>"
            + (f"<td class='procedure-col'>{html.escape(str(r['Procedure']))}</td>" if show_proc else "")
            + (f"<td class='attending-col'>{html.escape(str(r['Attending']))}</td>" if show_att else "")
            + f"<td class='comments-col'>{r['Comments_html']}</td>"
            f"</tr>"
        )
    _header_html = (
        "<th>Date</th>"
        + ("<th>Procedure</th>" if show_proc else "")
        + ("<th>Attending</th>" if show_att else "")
        + "<th>Comments</th>"
    )
    st.markdown(
        "<table class='comments-tbl'>"
        f"<thead><tr>{_header_html}</tr></thead>"
        f"<tbody>{_rows_html}</tbody></table>",
        unsafe_allow_html=True,
    )
    # Date, Procedure, and Attending default to the table's normal font
    # size (the CSS above just falls back to `inherit`) and Comments
    # absorbs the squeeze down to its own min-width first. Only if that
    # still isn't enough room does this shrink Date/Procedure/Attending
    # — together, to the same size as each other via one shared CSS
    # var — just enough to fit without wrapping.
    st.iframe(
        """
        <script>
        (function() {
            var doc = window.parent.document;
            var tables = doc.querySelectorAll('.comments-tbl');
            var table = tables[tables.length - 1];
            if (!table) return;
            function fit() {
                table.style.removeProperty('--cmts-sync-font');
                var container = table.parentElement;
                if (!container) return;
                var containerWidth = container.clientWidth;
                if (!containerWidth) return;
                var natural = table.scrollWidth;
                if (natural <= containerWidth) return;
                var dateCells = table.querySelectorAll('td.date-col');
                var procCells = table.querySelectorAll('td.procedure-col');
                var attCells = table.querySelectorAll('td.attending-col');
                if (!dateCells.length) return;
                function maxWidth(cells) {
                    var m = 0;
                    cells.forEach(function(c) { m = Math.max(m, c.scrollWidth); });
                    return m;
                }
                var threeW = maxWidth(dateCells) + maxWidth(procCells) + maxWidth(attCells);
                if (threeW <= 0) return;
                var otherW = natural - threeW;
                var availableForThree = containerWidth - otherW;
                var baseSize = parseFloat(window.getComputedStyle(dateCells[0]).fontSize);
                var ratio = Math.min(1, availableForThree / threeW) * 0.98;
                var newSize = Math.max(baseSize * ratio, 9);
                table.style.setProperty('--cmts-sync-font', newSize + 'px');
            }
            fit();
            window.parent.addEventListener('resize', fit);
            if (window.parent.ResizeObserver) {
                new window.parent.ResizeObserver(fit).observe(table.parentElement);
            }
        })();
        </script>
        """,
        height=1,
    )


def _build_resident_case_matrix(resident_email: str):
    """Every attending-confirmed case for one resident, joined down to
    per-step rating rows (self-assessments and steps left "Not Assessed"
    excluded, same as the resident's own Cumulative Dashboard).

    Returns (merged, steps_df, procs_map); `merged` is empty when there's
    no case or no meaningful rating yet — callers should treat that as
    "nothing to show" rather than distinguishing the two."""
    cases_df  = read_sheet_df(SHEET_CASES,  expected_cols=["case_id", "resident_email", "date",
                                                             "specialty_id", "procedure_id",
                                                             "attending_id", "notes",
                                                             "case_complexity", "overall_performance",
                                                             "assessment_type"])
    scores_df = read_sheet_df(SHEET_SCORES, expected_cols=["case_id", "step_id", "rating", "rating_num",
                                                             "case_complexity", "overall_performance"])
    steps_df  = read_sheet_df(SHEET_STEPS,  expected_cols=["step_id", "procedure_id", "step_order", "step_name"])
    procs_df  = read_sheet_df(SHEET_PROCEDURES, expected_cols=["procedure_id", "procedure_name", "specialty_id"])
    atnds_df  = _read_attendings_df()

    def _clean_id(val) -> str:
        s = str(val).strip()
        return s[:-2] if s.endswith(".0") else s

    atnds_lookup = {
        str(r.get("attending_id", "")): str(r.get("attending_name", ""))
        for _, r in atnds_df.iterrows()
    }
    procs_map = {
        str(r.get("procedure_id", "")): str(r.get("procedure_name", ""))
        for _, r in procs_df.iterrows()
    }

    resident_cases: dict = {}
    for _, row in cases_df.iterrows():
        if str(row.get("resident_email", "")).strip() != str(resident_email).strip():
            continue
        if str(row.get("assessment_type", "")).strip() == "Self-Assessment":
            continue
        cid = _clean_id(row.get("case_id", ""))
        if not cid or cid == "nan":
            continue
        aid = str(row.get("attending_id", ""))
        resident_cases[cid] = {
            "case_id":             cid,
            "date":                str(row.get("date", "")),
            "case_procedure_id":   str(row.get("procedure_id", "")),
            "attending_name":      attending_display_name(aid, atnds_lookup),
            "case_complexity":     row.get("case_complexity"),
            "overall_performance": row.get("overall_performance"),
        }

    if not resident_cases:
        return pd.DataFrame(), steps_df, procs_map

    steps_lookup: dict = {}
    for _, row in steps_df.iterrows():
        sid = str(row.get("step_id", "")).strip()
        if not sid or sid == "nan":
            continue
        steps_lookup[sid] = {
            "step_procedure_id": str(row.get("procedure_id", "")),
            "step_name":         str(row.get("step_name", "")),
            "step_order":        row.get("step_order", 0),
        }

    seen_case_step: set = set()
    merged_rows: list = []
    for _, row in scores_df.iterrows():
        cid = _clean_id(row.get("case_id", ""))
        if cid not in resident_cases:
            continue
        sid = str(row.get("step_id", "")).strip()
        if not sid or sid == "nan":
            continue
        key = (cid, sid)
        if key in seen_case_step:
            continue
        seen_case_step.add(key)
        step_meta = steps_lookup.get(sid, {})
        merged_rows.append({
            "case_id":             cid,
            "step_id":             sid,
            "rating":              str(row.get("rating", "")),
            "rating_num":          row.get("rating_num"),
            **resident_cases[cid],
            "step_procedure_id":   step_meta.get("step_procedure_id", ""),
            "step_name":           step_meta.get("step_name", ""),
            "step_order":          step_meta.get("step_order", 0),
        })

    _meaningful_case_ids = {r["case_id"] for r in merged_rows if r["rating"] != "Not Assessed"}
    merged_rows = [r for r in merged_rows if r["case_id"] in _meaningful_case_ids]

    if not merged_rows:
        return pd.DataFrame(), steps_df, procs_map

    merged = pd.DataFrame(merged_rows)
    if "case_procedure_id" not in merged.columns:
        merged["case_procedure_id"] = ""
    return merged, steps_df, procs_map


def _render_resident_heatmap(merged: pd.DataFrame, steps_df: pd.DataFrame, procs_map: dict,
                              selected_proc: str, filename_stub: str) -> None:
    """Render the progress heatmap + legends + Excel export for one
    resident's one procedure. `merged` is the resident's full case matrix
    from _build_resident_case_matrix (not yet filtered to a procedure) —
    this filters it to `selected_proc` itself."""
    proc_data = merged[merged["case_procedure_id"] == selected_proc].copy()
    if proc_data.empty:
        st.info("No assessment data yet for this procedure.")
        return

    ordered_steps = (
        steps_df[steps_df["procedure_id"] == selected_proc]
        .sort_values("step_order")["step_name"]
        .tolist()
    )

    def _fmt_step_hdr(name):
        """Column label shown above the heatmap: any "(...)" parenthetical
        is dropped, e.g. "Suture Placement (interrupted vs. running)"
        displays as just "Suture Placement". The full name is still used
        everywhere else (pivot table columns, dict keys) — this only
        changes what's shown in the header."""
        if not isinstance(name, str):
            return name
        return re.sub(r"\s*\([^)]*\)", "", name).strip()

    _step_display        = {s: _fmt_step_hdr(s) for s in ordered_steps}
    ordered_steps_display = [_step_display[s] for s in ordered_steps]

    pivot = proc_data.pivot_table(
        index=["date", "attending_name", "case_id", "overall_performance", "case_complexity"],
        columns="step_name",
        values="rating",
        aggfunc="first",
    ).reset_index()

    for step in ordered_steps:
        if step not in pivot.columns:
            pivot[step] = pd.NA

    pivot = pivot[["date", "attending_name", "case_id", "overall_performance", "case_complexity"] + ordered_steps]

    proc_display_name = procs_map.get(selected_proc, selected_proc)
    # One deliberate break point (see header_break_before): stays on one
    # line when it fits, and if it doesn't, wraps with "Progress Heatmap"
    # intact on the second line rather than splitting the procedure name
    # or "Progress"/"Heatmap" from each other.
    _heatmap_heading = header_break_before(f"{proc_display_name} —", "Progress Heatmap")
    st.markdown(
        f"### {_heatmap_heading}\n"
        "Most recent cases at the top. Zoom out to screenshot this grid. 📸"
    )
    st.caption("💡 Tip: To screenshot the full table — on mobile use print preview; on desktop use File > Print (Cmd+P / Ctrl+P), then adjust the scale percentage down until all columns fit on one page before screenshotting.")

    pivot_sorted = pivot.sort_values("date", ascending=False)

    _mr = {"date": "", "attending_name": "📌 Most Recent", "case_complexity": pd.NA, "overall_performance": pd.NA}
    for _s in ordered_steps:
        _vals = pivot_sorted[_s].dropna()
        _vals = _vals[_vals != "Not Assessed"]
        _mr[_s] = _vals.iloc[0] if not _vals.empty else pd.NA

    _best = {"date": "", "attending_name": "🏆 Best", "case_complexity": pd.NA, "overall_performance": pd.NA}
    for _s in ordered_steps:
        _vals = pivot_sorted[_s].dropna()
        if _vals.empty:
            _best[_s] = pd.NA
        else:
            _best[_s] = max(_vals.tolist(), key=lambda v: RATING_TO_NUM.get(v, -1))

    _summary_df = pd.DataFrame([_mr, _best])
    _meta_cols  = ["date", "attending_name", "overall_performance", "case_complexity"]

    display_df = pd.concat(
        [_summary_df[_meta_cols + ordered_steps],
         pivot_sorted.drop(columns=["case_id"])[_meta_cols + ordered_steps]],
        ignore_index=True,
    )

    display_df["date"] = display_df["date"].apply(fmt_date)

    display_df = display_df.rename(columns={
        "date":                "Date",
        "attending_name":      "Attending",
        "case_complexity":     "Case Complexity",
        "overall_performance": "Overall Performance",
        **_step_display,
    })
    all_cols = list(display_df.columns)

    display_df["Date"]      = display_df["Date"].fillna("")
    display_df["Attending"] = display_df["Attending"].fillna("")

    _rating_cols = [c for c in ordered_steps_display + ["Case Complexity", "Overall Performance"]
                    if c in display_df.columns]

    _orig_vals = {}
    for col in _rating_cols:
        _v = display_df[col].copy()
        if isinstance(_v, pd.DataFrame):
            _v = _v.iloc[:, 0]
        _orig_vals[col] = _v.reindex(display_df.index)

    for _c in _rating_cols:
        display_df[_c] = " "

    def _color_step(val):
        # A blank/NaN cell (no score row at all for this step on this
        # case) reads identically to an explicit "Not Assessed" rating —
        # same very light gray either way, since neither means anything
        # was actually observed.
        _is_na = val is None or (isinstance(val, float) and np.isnan(val)) or val == ""
        try:
            _is_na = _is_na or pd.isna(val)
        except (TypeError, ValueError):
            pass
        if _is_na:
            val = "Not Assessed"
        color = RATING_HEX.get(val, "")
        return f"background-color: {color}" if color else ""

    def _color_complexity(val):
        if pd.isna(val) or val == "":
            return ""
        return f"background-color: {COMPLEXITY_HEX.get(val, '')}"

    def _color_o_score(val):
        if not isinstance(val, str) or val == "":
            return ""
        key = val.split("-")[0].strip()
        return f"background-color: {O_SCORE_HEX.get(key, '')}"

    try:
        styled = display_df.style

        if ordered_steps_display:
            _safe_step_cols = [c for c in ordered_steps_display if c in _orig_vals]

            def _apply_step_colors(col):
                return [_color_step(v) for v in _orig_vals[col.name]]

            if _safe_step_cols:
                styled = styled.apply(_apply_step_colors, subset=_safe_step_cols, axis=0)

        def _apply_complexity_colors(col):
            return [_color_complexity(v) for v in _orig_vals["Case Complexity"]]

        def _apply_o_score_colors(col):
            return [_color_o_score(v) for v in _orig_vals["Overall Performance"]]

        # Every colored cell (step ratings, Case Complexity, Overall
        # Performance — all blanked to a single space, the color is the
        # only content) has a fixed width/height independent of this
        # table's actual rendered row height (a table cell's height is
        # only ever a minimum, and other cells in the same row can still
        # push it taller), tuned by eye. Case Complexity/Overall
        # Performance get their own, wider column than the step cells.
        _COLOR_CELL_HEIGHT_PX = 20
        _STEP_CELL_WIDTH_PX = 25
        _META_CELL_WIDTH_PX = 30  # Case Complexity / Overall Performance

        def _color_cell_props(width_px):
            return {
                "width": f"{width_px}px", "min-width": f"{width_px}px", "max-width": f"{width_px}px",
                "height": f"{_COLOR_CELL_HEIGHT_PX}px", "min-height": f"{_COLOR_CELL_HEIGHT_PX}px",
                "max-height": f"{_COLOR_CELL_HEIGHT_PX}px",
                "text-align": "center", "box-sizing": "border-box", "line-height": "1",
            }

        _STEP_CELL_PROPS = _color_cell_props(_STEP_CELL_WIDTH_PX)
        _META_CELL_PROPS = _color_cell_props(_META_CELL_WIDTH_PX)
        styled = (
            styled
            .apply(_apply_complexity_colors, subset=["Case Complexity"], axis=0)
            .apply(_apply_o_score_colors,    subset=["Overall Performance"], axis=0)
            .hide(axis="index")
            .set_properties(
                subset=["Date", "Attending"],
                # font-size/line-height set directly here (inline, same as
                # the colored cells' own props below) rather than relying
                # on the table-wide "th, td" rule to reach these two
                # columns — Date/Attending's text was still forcing the
                # row taller than the colored cells' own 25px square.
                **{"min-width": "120px", "white-space": "nowrap",
                   "font-size": "0.7rem", "line-height": "1"},
            )
            .set_properties(
                subset=["Case Complexity", "Overall Performance"],
                **_META_CELL_PROPS,
            )
        )
        if ordered_steps_display:
            styled = styled.set_properties(
                subset=ordered_steps_display,
                **_STEP_CELL_PROPS,
            )

        table_styles = [
            {"selector": "table",       "props": [("border-collapse", "collapse"), ("margin", "0 auto"),
                                                   ("border", "2px solid #555")]},
            {"selector": "th, td",      "props": [("border", "1px solid #bbb"),
                                                   ("padding", "4px"), ("font-size", "0.8rem"),
                                                   ("line-height", "1")]},
            {"selector": "th.col_heading", "props": [("text-align", "center"), ("vertical-align", "bottom"),
                                                       ("font-weight", "600")]},
            {"selector": "thead tr:last-child th", "props": [("border-bottom", "2px solid #555")]},
            {"selector": "tbody tr", "props": [("border-bottom", "1px solid #bbb")]},
            {"selector": "tbody tr:nth-child(1)", "props": [("border-bottom", "2px solid #555")]},
            {"selector": "tbody tr:nth-child(2)", "props": [("border-bottom", "2px solid #555")]},
            {"selector": "tbody tr:nth-child(1) td:nth-child(1)",
             "props": [("border-right", "none")]},
            {"selector": "tbody tr:nth-child(2) td:nth-child(1)",
             "props": [("border-right", "none")]},
            {"selector": "tbody tr:nth-child(1) td:nth-child(2)",
             "props": [("text-align", "right"), ("font-weight", "600"),
                       ("padding-right", "6px"), ("border-left", "none")]},
            {"selector": "tbody tr:nth-child(2) td:nth-child(2)",
             "props": [("text-align", "right"), ("font-weight", "600"),
                       ("padding-right", "6px"), ("border-left", "none")]},
        ]
        _vheader_cols  = [c for c in all_cols
                           if c in ordered_steps_display or c in ("Case Complexity", "Overall Performance")]

        # Rotated column headers: forced to one line (no wrapping, no
        # shrink-to-fit) with no cap on the header row's height — a long
        # label just makes that row taller instead of wrapping or
        # shrinking. Revisit if very long labels end up making the
        # header row uncomfortably tall.
        for idx, col_name in enumerate(all_cols):
            if col_name in _vheader_cols:
                _hdr_width_px = (
                    _META_CELL_WIDTH_PX if col_name in ("Case Complexity", "Overall Performance")
                    else _STEP_CELL_WIDTH_PX
                )
                table_styles.append({
                    "selector": f"th.col_heading.level0.col{idx}",
                    "props": [
                        ("writing-mode", "vertical-rl"),
                        ("transform", "rotate(180deg)"),
                        ("vertical-align", "bottom"),
                        ("text-align", "left"),
                        ("padding", "4px 2px"),
                        ("width", f"{_hdr_width_px}px"),
                        ("min-width", f"{_hdr_width_px}px"),
                        ("max-width", f"{_hdr_width_px}px"),
                        ("white-space", "nowrap"),
                        ("font-size", "0.75rem"),
                    ],
                })

        table_styles.append({
            "selector": "th.col_heading .pp-vhdr-inner",
            "props": [
                ("display", "flex"),
                ("align-items", "center"),
                ("justify-content", "flex-start"),
                ("width", "100%"),
                ("height", "100%"),
            ],
        })

        def _wrap_vheader_labels(html_str, vheader_indices):
            if not vheader_indices:
                return html_str
            _col_idx_re = re.compile(r"\bcol(\d+)\b")

            def _wrap(m):
                open_tag, inner, close_tag = m.group(1), m.group(2), m.group(3)
                if "col_heading" not in open_tag:
                    return m.group(0)
                idx_match = _col_idx_re.search(open_tag)
                if not idx_match or int(idx_match.group(1)) not in vheader_indices:
                    return m.group(0)
                return f'{open_tag}<div class="pp-vhdr-inner">{inner}</div>{close_tag}'

            return re.sub(r"(<th\b[^>]*>)(.*?)(</th>)", _wrap, html_str, flags=re.DOTALL)

        styled = styled.set_table_styles(table_styles)
        _vheader_idx = {idx for idx, c in enumerate(all_cols) if c in _vheader_cols}
        st.markdown(_wrap_vheader_labels(styled.to_html(), _vheader_idx), unsafe_allow_html=True)

    except Exception as _heatmap_err:
        st.warning(
            f"⚠️ Could not render the heatmap for this procedure: {_heatmap_err}\n\n"
            "Please try a different procedure, or contact your program coordinator."
        )

    def _swatch(color, label, border=""):
        _bdr = f"border:{border};" if border else ""
        return (
            f'<span class="legend-item">'
            f'<span class="legend-swatch" style="background-color:{color};{_bdr}"></span>{label}'
            f'</span>'
        )

    st.markdown("#### Ratings Legend")
    _rating_legend_html = _swatch("#E0E0E0", "Never Attempted")
    _rating_legend_html += "".join(
        _swatch(v, k, "1px solid #aaa" if k == "Not Assessed" else "")
        for k, v in RATING_HEX.items()
    )
    st.markdown('<div class="legend-row">' + _rating_legend_html + "</div>", unsafe_allow_html=True)

    st.markdown("#### Case Complexity")
    st.markdown(
        '<div class="legend-row">' +
        "".join(_swatch(v, k) for k, v in COMPLEXITY_HEX.items()) +
        "</div>",
        unsafe_allow_html=True,
    )

    st.markdown("---")
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pivot_excel = pivot.copy()
        pivot_excel["date"] = pivot_excel["date"].apply(fmt_date)
        pivot_excel = pivot_excel.rename(columns={
            "date":                "Date",
            "attending_name":      "Attending",
            "case_id":             "Case ID",
            "case_complexity":     "Case Complexity",
            "overall_performance": "Overall Performance",
        })
        pivot_excel.to_excel(writer, index=False, sheet_name="Cumulative")
        ws_xl = writer.sheets["Cumulative"]
        from openpyxl.styles import PatternFill, Font

        step_fill_map = {k: v.lstrip("#") for k, v in RATING_HEX.items() if k not in ("Not Assessed",)}
        step_fill_map["Not Assessed"] = "E0E0E0"

        start_col = 6
        for xl_row in ws_xl.iter_rows(
            min_row=2, max_row=ws_xl.max_row,
            min_col=start_col, max_col=5 + len(ordered_steps),
        ):
            for cell in xl_row:
                val = cell.value
                if val in step_fill_map:
                    cell.fill = PatternFill(
                        start_color=step_fill_map[val],
                        end_color=step_fill_map[val],
                        fill_type="solid",
                    )
                    cell.font = Font(color="FFFFFF" if val in ("Not Yet", "Auto") else "000000")

    st.download_button(
        label=f"📥 Download Excel — {procs_map.get(selected_proc, selected_proc)}",
        data=output.getvalue(),
        file_name=f"{filename_stub}_{selected_proc}_cumulative.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"dl_heatmap_{filename_stub}_{selected_proc}",
    )


# ─────────────────────────────────────────────
# PAGE HEADER HELPER
# ─────────────────────────────────────────────
def _header_max(text: str) -> float:
    """Font-size ceiling (rem) for a header, tuned to text length. There is
    deliberately no floor — on skinny windows the font should keep shrinking
    (down to whatever -webkit-line-clamp: 2 and ellipsis allow) rather than
    being held at a minimum size and forced to wrap or truncate."""
    length = len(text)
    if length <= 20:
        return 2.75
    elif length <= 35:
        return 2.3
    elif length <= 55:
        return 1.9
    else:
        return 1.6


def _protect_from_wrapping(text: str) -> str:
    """Replace every space and hyphen with its non-breaking Unicode
    counterpart, so nothing inside `text` offers the browser a place
    to wrap — a plain hyphen is a valid line-break opportunity on its
    own by default (independent of spaces), which let a procedure name
    like "Robotic-Assisted ..." break there instead of only at a
    forced-break helper's designated point."""
    return text.replace(" ", " ").replace("-", "‑")


def header_break_before(prefix: str, suffix: str) -> str:
    """Join "{prefix} {suffix}", protecting every space and hyphen in
    each part from wrapping except the single regular space between
    them — so if page_header()'s mobile 2-line allowance (see its own
    fit script) ever does need to wrap this header, the break can only
    land right at that boundary, never mid-word within either part."""
    return f"{_protect_from_wrapping(prefix)} {_protect_from_wrapping(suffix)}"


def suppress_picker_keyboards() -> None:
    """Set inputmode="none" on every st.selectbox and st.date_input
    input on the currently rendered page, so tapping one to open its
    dropdown/calendar doesn't also pop up the on-screen keyboard on
    mobile — they're pick-from-a-list-or-calendar controls, not
    free-text fields anyone needs to type into. Global (not scoped to
    specific widget keys) so it covers every such widget on every page,
    including ones added later, without each needing its own key and
    call site. A MutationObserver re-applies it whenever Streamlit
    re-renders an input (e.g. after a selection), since a plain
    one-time pass would only catch whatever's in the DOM at that
    instant. Called once, after every page render (see the bottom of
    this file), same as fit_all_button_labels().

    The date input also gets readonly: inputmode="none" alone wasn't
    enough to stop it opening the keyboard — confirmed it renders as a
    plain type="text" input, where mobile browser support for
    respecting inputmode="none" is inconsistent. readonly blocks any
    on-screen keyboard unconditionally, and (verified directly) doesn't
    stop it from being clicked/tapped to open the calendar popup, which
    is the only way this field is meant to be filled in anyway.
    Selectboxes don't get readonly: they support typing to search/
    filter their own option list, which readonly would break.

    Confirmed on a real Android phone (both Chrome and Edge — same
    Chromium engine, so consistent with an Android/Chromium-level
    behavior rather than a browser-specific quirk) that none of that
    stopped the keyboard, which stayed up until manually dismissed.
    Two more layers were tried and kept as defense-in-depth, but the
    one that actually solves it for a real touch is the last one below:

    1. BaseWeb (the underlying component library) can reset an input's
       own DOM attributes on a React re-render faster than the
       childList/subtree observer below reacts to it — that observer
       only fires on nodes being added/removed, not on an existing
       node's attributes changing. Each date input additionally gets
       its own dedicated attribute-level observer, which reacts to
       exactly that case.

    2. A focus listener that blurs the input ~50ms after it's focused.
       Confirmed directly that the calendar popup opens on focus but
       doesn't close again on blur (its open state isn't tied to
       staying focused), so this dismisses whatever keyboard attempt
       is in flight without dismissing the popup — an immediate
       (0ms) blur was tried first and closed the popup too, so the
       delay is deliberate: long enough for the popup's own
       open-on-focus effect to have already run.

    3. The actual fix: a real touch landing directly on an editable
       input is what triggers a mobile OS keyboard, regardless of
       readonly/inputmode/blur timing — so stop a real touch from ever
       reaching the input at all. The input gets pointer-events: none,
       and an invisible same-sized overlay div sits on top of it inside
       its immediate wrapper (BaseWeb's own [data-baseweb="base-input"]
       div, sized to match the input already, so the overlay's CSS
       inset:0 tracks it through any resize with no JS recalculation
       needed); tapping the overlay calls el.focus() programmatically
       instead. A script-triggered focus that didn't originate from a
       direct touch on that specific element is the one thing that
       reliably does NOT bring up the keyboard on Android Chrome/Edge —
       and (verified directly) still opens the calendar popup and
       supports picking a date from it, same as a real click always
       did."""
    st.iframe(
        """
        <script>
        (function() {
            var doc = window.parent.document;
            function apply() {
                doc.querySelectorAll('[data-testid="stSelectbox"] input').forEach(function(el) {
                    el.setAttribute('inputmode', 'none');
                });
                doc.querySelectorAll('[data-testid="stDateInput"] input').forEach(function(el) {
                    el.setAttribute('inputmode', 'none');
                    el.setAttribute('readonly', 'readonly');
                    if (el.__ppKeyboardGuard) return;
                    el.__ppKeyboardGuard = true;
                    if (window.parent.MutationObserver) {
                        new window.parent.MutationObserver(function() {
                            if (el.getAttribute('inputmode') !== 'none') {
                                el.setAttribute('inputmode', 'none');
                            }
                            if (!el.hasAttribute('readonly')) {
                                el.setAttribute('readonly', 'readonly');
                            }
                        }).observe(el, {attributes: true, attributeFilter: ['inputmode', 'readonly']});
                    }
                    el.addEventListener('focus', function() {
                        setTimeout(function() { el.blur(); }, 50);
                    });
                    var wrapper = el.closest('[data-baseweb="base-input"]') || el.parentElement;
                    if (window.parent.getComputedStyle(wrapper).position === 'static') {
                        wrapper.style.position = 'relative';
                    }
                    el.style.pointerEvents = 'none';
                    var overlay = doc.createElement('div');
                    overlay.style.position = 'absolute';
                    overlay.style.inset = '0';
                    overlay.style.zIndex = '5';
                    overlay.style.cursor = 'pointer';
                    overlay.addEventListener('click', function(e) {
                        e.preventDefault();
                        el.focus();
                    });
                    wrapper.appendChild(overlay);
                });
            }
            apply();
            if (window.parent.MutationObserver) {
                new window.parent.MutationObserver(apply).observe(doc.body, {
                    childList: true, subtree: true
                });
            }
        })();
        </script>
        """,
        height=1,
    )


def page_header(text: str, tier_text: str | None = None) -> None:
    """Render a page's main H1 header, then measure its actual rendered
    width in the browser and scale the font to exactly fill the
    container — no leftover right-hand margin — while never exceeding
    this length tier's max and never wrapping past two lines (a CSS
    safety net in case the measurement can't run, e.g. scripts disabled).
    A per-character cqw estimate can't do this precisely: real text
    width depends on which letters are in it, not just how many, so a
    constant safe enough to avoid ever wrapping always left a visible
    gap for shorter/narrower strings. Measuring the actual rendered
    width removes that guesswork entirely. Field labels, dropdown values,
    Step-Level Ratings, and Improve/How no longer scale off this header's
    size (see --pp-substep-font's own static definition) — a long title
    forcing the header down to avoid wrapping used to shrink everything
    else on the page right along with it.

    tier_text: text to base _header_max()'s length tier on, if different
    from what's actually displayed — e.g. a variable suffix (a resident's
    name) that shouldn't itself push the header into a smaller ceiling
    tier just for being long. The real fit still measures the full
    displayed text's actual rendered width, so it still shrinks further
    than that ceiling if the full text doesn't fit."""
    max_rem = _header_max(tier_text if tier_text is not None else text)
    escaped = html.escape(text)
    st.markdown(
        f'<div class="pp-page-header-wrap"><h1 class="pp-page-header">'
        f'{escaped}</h1></div>',
        unsafe_allow_html=True,
    )
    st.iframe(
        f"""
        <script>
        (function() {{
            var doc = window.parent.document;
            var wraps = doc.querySelectorAll('.pp-page-header-wrap');
            var wrap = wraps[wraps.length - 1];
            if (!wrap) return;
            var el = wrap.querySelector('.pp-page-header');
            if (!el) return;
            var maxPx = {max_rem} * 16;
            var nbsp = '\\u00a0';
            // Measures a string's rendered single-line width at a given
            // font size via a detached, invisible probe — never touches
            // el itself, so its real DOM (Streamlit wraps header text in
            // an anchor-link span) is never disturbed.
            function measureWidth(str, fontPx) {{
                var probe = doc.createElement('span');
                probe.style.position = 'absolute';
                probe.style.visibility = 'hidden';
                probe.style.whiteSpace = 'nowrap';
                probe.style.fontSize = fontPx + 'px';
                var computed = window.parent.getComputedStyle(el);
                probe.style.fontFamily = computed.fontFamily;
                probe.style.fontWeight = computed.fontWeight;
                probe.textContent = str;
                doc.body.appendChild(probe);
                var w = probe.scrollWidth;
                doc.body.removeChild(probe);
                return w;
            }}
            function fit() {{
                var containerWidth = wrap.clientWidth;
                if (!containerWidth) return;
                var fullText = el.textContent;
                // header_break_before() (see its Python definition)
                // builds some headers with
                // exactly one regular, breakable space and non-breaking
                // spaces (nbsp) everywhere else, marking one deliberate
                // wrap point. On a narrow (mobile) screen, measure each
                // side of that point separately and take the wider one —
                // rather than assuming (as an earlier version of this
                // script did) that the text splits into two *even*
                // halves, which let a lopsided split (e.g. a long
                // procedure name, short resident name) overflow its line
                // and trigger the CSS's overflow-wrap: break-word
                // mid-word. Headers without such a point (or on desktop)
                // measure as one line, same as always.
                var mobile = window.parent.innerWidth <= 600;
                var breakIdx = mobile && fullText.indexOf(nbsp) > -1
                    ? fullText.indexOf(' ') : -1;
                var widest = breakIdx > -1
                    ? Math.max(
                        measureWidth(fullText.slice(0, breakIdx), maxPx),
                        measureWidth(fullText.slice(breakIdx + 1), maxPx)
                      )
                    : measureWidth(fullText, maxPx);
                // The two-segment (breakIdx > -1) case approximates two
                // separate wrapped lines from single-line nowrap probe
                // measurements of each segment — each segment (kept
                // unbreakable internally via nbsp) still has to survive
                // the browser's own multi-line layout afterward, which
                // can round a hair differently than the probe. An
                // unusually long unbroken segment (e.g. a long combined
                // first+last name) could then overflow its line by a
                // pixel or two and get pushed onto a clipped 3rd line by
                // the CSS's -webkit-line-clamp safety net. The plain
                // single-line case doesn't have that extra layout step,
                // so it keeps the tighter margin that was already tuned
                // to fill the header's width precisely.
                var safety = breakIdx > -1 ? 0.9 : 0.96;
                var finalPx = widest <= containerWidth
                    ? maxPx
                    : Math.max(1, maxPx * (containerWidth / widest) * safety);
                el.style.fontSize = finalPx + 'px';
            }}
            fit();
            window.parent.addEventListener('resize', fit);
            // Also watch the container itself: Streamlit's own column/
            // layout reflow can change its width slightly after this
            // script's first run, with no window 'resize' event to catch
            // it — a plain fixed-width guess would miss that follow-up.
            if (window.parent.ResizeObserver) {{
                new window.parent.ResizeObserver(fit).observe(wrap);
            }}
        }})();
        </script>
        """,
        height=1,
    )


def assessment_instructions_note() -> None:
    """Info box explaining the assessment form's three main sections —
    shown under the page header, before the first divider, on all three
    assessment-filling pages (Assess Together, Self-Assess, and the
    attending's pre-filled/blank forms)."""
    st.info(
        "There are 3 main sections. Fill out as much or as little as you are able.\n\n"
        "1. Short Form: Improve this / Do this\n"
        "2. Step-Level Ratings of a Case or Skills\n"
        "3. Free Form: Development/Improvement/Feed-Forward"
    )


def copy_link_button(link: str, key: str) -> None:
    """Big, obvious "Copy Link" button. st.code()'s own built-in copy
    icon is small, only shows on hover, and on mobile needs a first tap
    just to reveal it — easy to miss entirely.

    Rendered via st.iframe rather than st.markdown(unsafe_allow_html=True):
    confirmed empirically that Streamlit strips onclick (and presumably
    any other inline event-handler attribute) from markdown HTML even
    with unsafe_allow_html=True — the button still renders, just inert,
    with no error or warning. st.iframe's content isn't run through that
    sanitizer at all, and — also confirmed empirically (an actual OS
    clipboard write was observed after a click, matching the intended
    text) — the Clipboard API works fine from inside it. Falls back to
    the older execCommand('copy') approach (via a temporary off-screen
    textarea) if navigator.clipboard isn't available."""
    safe_link = json.dumps(link)
    st.iframe(
        f"""
        <!DOCTYPE html>
        <html><head><style>
        body {{ margin: 0; font-family: "Source Sans Pro", sans-serif; }}
        button {{
            width: 100%;
            box-sizing: border-box;
            padding: 0.6rem 1rem;
            font-size: 1rem;
            font-weight: 600;
            border-radius: 8px;
            border: 3px solid #FF4B4B;
            background: #FFFFFF;
            color: #000000;
            cursor: pointer;
        }}
        button:hover {{
            background: #FFF0F0;
            border-color: #E63946;
        }}
        </style></head>
        <body>
        <button id="{key}">📋 Copy Link</button>
        <script>
        var text = {safe_link};
        var btn = document.getElementById("{key}");
        btn.addEventListener("click", function() {{
            function done(ok) {{
                btn.textContent = ok ? "✅ Copied!" : "⚠️ Copy failed — select manually below";
                setTimeout(function() {{ btn.textContent = "📋 Copy Link"; }}, 1800);
            }}
            function fallback() {{
                var ta = document.createElement("textarea");
                ta.value = text;
                ta.style.position = "fixed";
                ta.style.opacity = "0";
                document.body.appendChild(ta);
                ta.focus();
                ta.select();
                var ok = false;
                try {{ ok = document.execCommand("copy"); }} catch (e) {{}}
                document.body.removeChild(ta);
                done(ok);
            }}
            if (navigator.clipboard && navigator.clipboard.writeText) {{
                navigator.clipboard.writeText(text).then(function() {{ done(true); }}, fallback);
            }} else {{
                fallback();
            }}
        }});
        </script>
        </body></html>
        """,
        height=54,
    )


def mobile_tip(text: str) -> None:
    """Render the "On mobile: ..." tip, then shrink its font (down from the
    CSS-defined base size in the .st-key-mobile_tip rule) just enough that
    the label never wraps past one line — same measure-and-fit approach as
    page_header. white-space:nowrap + text-overflow:ellipsis in the CSS is
    the fallback if the measurement can't run (scripts disabled)."""
    with st.container(key="mobile_tip"):
        st.info(text)
    st.iframe(
        """
        <script>
        (function() {
            var doc = window.parent.document;
            var ps = doc.querySelectorAll('.st-key-mobile_tip [data-testid="stAlertContainer"] p');
            var el = ps[ps.length - 1];
            if (!el) return;
            // el's own clientWidth is unreliable here: with white-space:nowrap
            // forced (from the CSS fallback), the <p> won't shrink below its
            // own unwrapped content size, so el.clientWidth just reports that
            // same overflowing size instead of the space actually available.
            // Streamlit's stMarkdownContainer wrapper div, one level up,
            // already gets an explicit width (accounting for the icon) that
            // isn't affected by the <p>'s own sizing — measure against that.
            var avail = el.parentElement;
            function fit() {
                el.style.fontSize = '';
                var availWidth = avail.clientWidth;
                if (!availWidth) return;
                var natural = el.scrollWidth;
                if (natural > availWidth) {
                    var baseSize = parseFloat(window.getComputedStyle(el).fontSize);
                    el.style.fontSize = Math.max(1, baseSize * (availWidth / natural) * 0.98) + 'px';
                }
            }
            fit();
            window.parent.addEventListener('resize', fit);
            if (window.parent.ResizeObserver) {
                new window.parent.ResizeObserver(fit).observe(avail);
            }
        })();
        </script>
        """,
        height=1,
    )


def fit_all_button_labels() -> None:
    """Shrink any button's label just enough that it never wraps past one
    line — same measure-and-fit approach as page_header/mobile_tip, just
    applied to every button on the currently rendered page at once.
    white-space:nowrap in the global CSS is the fallback if the
    measurement can't run (scripts disabled) — it clips with an ellipsis
    instead of wrapping to a second line."""
    st.iframe(
        """
        <script>
        (function() {
            var doc = window.parent.document;
            function fit(p) {
                var btn = p.closest('button');
                if (!btn) return;
                p.style.fontSize = '';
                // Measure against the <p>'s own box, not the button's —
                // the button is wider than its label (padding), so
                // comparing against btn.clientWidth let text through
                // that still overflowed the <p>'s own narrower box and
                // got silently clipped by its ellipsis fallback instead
                // of actually being shrunk to fit.
                var availWidth = p.clientWidth;
                if (!availWidth) return;
                var natural = p.scrollWidth;
                if (natural > availWidth) {
                    var baseSize = parseFloat(window.getComputedStyle(p).fontSize);
                    var target = baseSize * (availWidth / natural) * 0.9;
                    p.style.fontSize = Math.max(target, baseSize * 0.55, 9) + 'px';
                }
            }
            function fitAll() {
                doc.querySelectorAll('button p').forEach(fit);
            }
            fitAll();
            window.parent.addEventListener('resize', fitAll);
            if (window.parent.ResizeObserver) {
                new window.parent.ResizeObserver(fitAll).observe(doc.body);
            }
        })();
        </script>
        """,
        height=1,
    )


def sync_improve_how_label_width() -> None:
    """Measure the "In order to improve this:" and "Do this:" labels'
    natural (unwrapped) widths and set --pp-improve-label-width to the
    wider of the two — same measure-and-fit approach as page_header/
    mobile_tip, but syncing a shared width across two separate st.columns()
    rows instead of a font-size. The CSS rule using that var then gives
    both label columns that same width, so each label sits snug against
    its own text box and the two text boxes' left edges line up between
    the two rows, regardless of which row's label text is longer. Call
    this once, right after rendering both rows."""
    st.iframe(
        """
        <script>
        (function() {
            var doc = window.parent.document;
            var containers = doc.querySelectorAll('.st-key-assess_improve_how');
            var container = containers[containers.length - 1];
            if (!container) return;
            var labels = container.querySelectorAll('[data-testid="stMarkdownContainer"] p');
            function fit() {
                var maxWidth = 0;
                labels.forEach(function(p) {
                    var prevDisplay = p.style.display;
                    var prevWhiteSpace = p.style.whiteSpace;
                    p.style.display = 'inline-block';
                    p.style.whiteSpace = 'nowrap';
                    maxWidth = Math.max(maxWidth, p.scrollWidth);
                    p.style.display = prevDisplay;
                    p.style.whiteSpace = prevWhiteSpace;
                });
                if (maxWidth > 0) {
                    container.style.setProperty('--pp-improve-label-width', (maxWidth + 3) + 'px');
                }
            }
            fit();
            window.parent.addEventListener('resize', fit);
            if (window.parent.ResizeObserver) {
                new window.parent.ResizeObserver(fit).observe(container);
            }
        })();
        </script>
        """,
        height=1,
    )


# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
st.sidebar.title("🩺 Procedure Passport")

_is_attending_role = st.session_state.get("role") == "attending"
# When an attending is logged in, "resident" holds whichever resident they're
# currently assessing (same repurposing the anonymous magic-link flow already
# does) — not their own identity, so it must never drive the resident-login
# sidebar below.
_logged_in = None if _is_attending_role else st.session_state.get("resident")
_attending_logged_in = st.session_state.get("attending_login_email") if _is_attending_role else None

if _logged_in in ADMINS:
    if st.sidebar.button("⚙️ Admin Panel"):
        go_to("admin")

if _logged_in and st.session_state["page"] not in ("login", "attending_assessment", "attending_confirmation"):
    st.sidebar.markdown(f"👤 **{st.session_state.get('resident_name', '')}**")
    st.sidebar.markdown(f"_{_logged_in}_")
    st.sidebar.markdown("---")
    if st.sidebar.button("🚪 Logout", key="sb_logout_resident"):
        for _k in list(st.session_state.keys()):
            del st.session_state[_k]
        st.cache_data.clear()
        st.rerun()

# ── Sidebar nav shortcuts (shown when logged in on relevant pages) ──
if _logged_in and st.session_state["page"] not in ("login", "attending_assessment", "attending_confirmation"):
    st.sidebar.markdown("---")
    if st.sidebar.button("➕ Start Assessment", key="sb_start"):
        st.session_state["page"] = "start"
        st.rerun()
    if st.sidebar.button("📊 Cumulative Dashboard", key="sb_cumulative"):
        st.session_state["page"] = "cumulative"
        st.rerun()
    if st.sidebar.button("💬 Comments Dashboard", key="sb_comments"):
        st.session_state["page"] = "comments"
        st.rerun()
    if st.sidebar.button("🏠 Back to Home", key="sb_home"):
        st.session_state["page"] = "home"
        st.rerun()

# ── Attending-account sidebar (own login, separate from the anonymous
# magic-link flow above) ──
if _attending_logged_in:
    st.sidebar.markdown(f"👤 **{st.session_state.get('attending_login_name', '')}**")
    st.sidebar.markdown(f"_{_attending_logged_in}_")
    st.sidebar.markdown("---")
    if st.sidebar.button("🚪 Logout", key="sb_logout_attending"):
        for _k in list(st.session_state.keys()):
            del st.session_state[_k]
        st.cache_data.clear()
        st.rerun()
    st.sidebar.markdown("---")
    if st.sidebar.button("➕ Start Assessment", key="sb_att_start"):
        st.session_state["page"] = "attending_start"
        st.rerun()
    if st.sidebar.button("📊 Resident Dashboard", key="sb_att_dashboard"):
        st.session_state["page"] = "attending_resident_dashboard"
        st.rerun()
    if st.sidebar.button("🏠 Back to Home", key="sb_att_home"):
        st.session_state["page"] = "attending_home"
        st.rerun()

# ── Sidebar rating legend (shown only on relevant pages) ──
if st.session_state.get("page") in (
    "start", "assessment", "dashboard", "cumulative", "attending_assessment",
    "attending_start", "attending_resident_dashboard",
):
    st.sidebar.markdown("---")
    st.sidebar.markdown("**Rating Scale**")
    _LEGEND_ITEMS = [
        ("Never Attempted","#E0E0E0", ""),
        ("Not Assessed",   "#FAFAFA", "1px solid #aaa"),
        ("Shown/Told",     "#9E9E9E", ""),
        ("Not Yet",        "#378ADD", ""),
        ("Steer",          "#FF944D", ""),
        ("Prompt",         "#FFD633", ""),
        ("Back up",        "#99E699", ""),
        ("Auto",           "#33CC33", ""),
    ]
    for _label, _color, _border in _LEGEND_ITEMS:
        _border_css = f"border:{_border};" if _border else ""
        st.sidebar.markdown(
            f'<span style="display:inline-block;width:13px;height:13px;'
            f'background:{_color};{_border_css}border-radius:2px;'
            f'margin-right:6px;vertical-align:middle;"></span>{_label}',
            unsafe_allow_html=True,
        )

# ─────────────────────────────────────────────
# SHARED CSS
# ─────────────────────────────────────────────
st.markdown(
    """
<style>
/* Every button's label stays on one line — fit_all_button_labels()
   shrinks the font to make it fit; this is the no-JS fallback (clips
   with an ellipsis instead of wrapping to a second line). */
button p {
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
/* Card-style sections */
.pp-card {
    background: var(--secondary-background-color);
    border-radius: 10px;
    padding: 1.2rem 1.5rem;
    margin-bottom: 1rem;
}
/* Pill badge */
.pp-badge {
    display: inline-block;
    border-radius: 12px;
    padding: 2px 10px;
    font-size: 0.82rem;
    font-weight: 600;
    margin: 2px;
}
/* Legend row */
.legend-row {
    display: flex;
    gap: 1rem;
    flex-wrap: wrap;
    align-items: center;
    margin-bottom: 0.5rem;
}
.legend-item {
    display: inline-flex;
    align-items: center;
    gap: 0.35rem;
    font-size: 0.85rem;
}
.legend-swatch {
    width: 14px;
    height: 14px;
    border-radius: 3px;
    border: 1px solid var(--secondary-background-color);
    display: inline-block;
}
/* Home page cards: keep the three action buttons vertically aligned
   even when title/description text wraps to different heights. */
.st-key-home_cards [data-testid="stColumn"] > [data-testid="stVerticalBlock"] {
    display: flex;
    flex-direction: column;
    height: 100%;
    gap: 0.3rem;
}
.st-key-home_cards [data-testid="stElementContainer"]:has([data-testid="stButton"]) {
    margin-top: auto;
}
.st-key-home_cards h3 [data-testid="stHeaderActionElements"] {
    display: none;
}
.st-key-home_cards [data-testid="stElementContainer"]:has(h3) {
    height: 5rem;
    overflow: hidden;
    container-type: inline-size;
}
.st-key-home_cards [data-testid="stElementContainer"]:has(h3) h3 {
    padding-top: 0.3rem;
    padding-bottom: 0.2rem;
}
.st-key-home_cards [data-testid="stElementContainer"]:has(h3) h3 > span:first-child {
    display: -webkit-box;
    -webkit-line-clamp: 2;
    -webkit-box-orient: vertical;
    overflow: hidden;
    font-size: clamp(0.95rem, 12cqw, 1.75rem);
    line-height: 1.25;
}
.st-key-home_cards [data-testid="stElementContainer"]:has(> [data-testid="stMarkdown"] p) {
    height: 2.4rem;
    overflow: hidden;
    container-type: inline-size;
}
.st-key-home_cards [data-testid="stElementContainer"]:has(> [data-testid="stMarkdown"] p) p {
    display: -webkit-box;
    -webkit-line-clamp: 2;
    -webkit-box-orient: vertical;
    overflow: hidden;
    margin: 0;
    line-height: 1.3;
    font-size: clamp(0.7rem, 8.5cqw, 0.875rem);
}
/* Step-Level Ratings expander: label styled like a smaller page header.
   --pp-substep-font sets its size — a fixed value (see its own
   definition below), independent of the main header's size, so a long
   procedure/resident name that shrinks the header doesn't shrink this
   too. white-space: normal overrides Streamlit's own default (nowrap +
   ellipsis) for expander summary labels, which is otherwise sized for
   the typical short, single-line case — header_break_before() keeps
   "Step-Level Ratings for" and the procedure name each as their own
   unbreakable unit, with the one regular breakable space landing right
   after "for": one line whenever it fits, and if not, the wrap lands
   there rather than mid-phrase or partway through the name.
   overflow-wrap: break-word is a safety net for a name too long to
   fit even on its own full line. */
.st-key-step_ratings_expander_resident summary [data-testid="stMarkdownContainer"] p,
.st-key-step_ratings_expander_attending summary [data-testid="stMarkdownContainer"] p {
    font-size: var(--pp-substep-font, 1.3125rem);
    font-weight: 600;
    white-space: normal !important;
    overflow-wrap: break-word;
}
/* "Click to Expand" hint on its own line under that label, at 3/4 of
   its size — a generated ::after (rather than a second line of real
   text) since st.expander's label is a single inline markdown string. */
.st-key-step_ratings_expander_resident summary [data-testid="stMarkdownContainer"] p::after,
.st-key-step_ratings_expander_attending summary [data-testid="stMarkdownContainer"] p::after {
    content: "Click to Expand";
    display: block;
    font-size: calc(var(--pp-substep-font, 1.3125rem) * 0.75);
    font-weight: normal;
}
/* Green border while the Step-Level Ratings expander is collapsed, to
   draw the eye to it; gone once it's opened (the native <details>
   element's own "open" attribute drives this, no JS needed here). */
.st-key-step_ratings_expander_resident details:not([open]),
.st-key-step_ratings_expander_attending details:not([open]) {
    border: 2px solid #2E7D32 !important;
}
/* Main page headers: font-size is set by page_header()'s injected script
   after measuring the real rendered text width, so it exactly fills the
   container. This is the CSS-only fallback/safety net (script disabled,
   or before the script's first paint): still capped at two lines. */
.pp-page-header {
    display: -webkit-box;
    -webkit-line-clamp: 2;
    -webkit-box-orient: vertical;
    overflow: hidden;
    line-height: 1.25;
    overflow-wrap: break-word;
    text-overflow: ellipsis;
}
/* Attending pre-filled-form notice: font-size is set by its own
   injected script (see the "if _draft:" block on the attending
   assessment page), which tries the whole notice on one line first,
   only shrinking down to a floor before allowing the deliberate
   two-line break header_break_before() marks. This is the CSS-only
   fallback/safety net (script disabled, or before its first paint):
   still capped at two lines. */
.pp-prefill-notice-text {
    display: -webkit-box;
    -webkit-line-clamp: 2;
    -webkit-box-orient: vertical;
    overflow: hidden;
    line-height: 1.3;
    overflow-wrap: break-word;
}
/* --pp-substep-font sizes a whole hierarchy of form text (field labels,
   dropdown values, Step-Level Ratings, Improve/How, the mobile tip).
   It used to be tied to the page header's own live-measured size, so
   the page scaled as one unit with window width — but a long procedure/
   resident name could force the header down quite small to avoid
   wrapping (by design, see page_header() — no floor there), and
   everything tied to it shrank right along with it, well past
   comfortable reading size, on desktop as much as mobile. Fixed at all
   widths now, independent of the header entirely. */
:root {
    --pp-substep-font: 1.3125rem;
}
/* Assessment page top nav: Streamlit stacks columns onto separate rows
   below a width breakpoint (each stColumn gets min-width: ~100%). Force
   the Back/Home buttons to stay side by side at any width instead. Columns
   don't shrink past their button's natural (nowrap) width, so on very
   narrow screens the row scrolls horizontally rather than the two
   buttons shrinking into each other and overlapping. */
.st-key-assess_top_nav [data-testid="stHorizontalBlock"] {
    flex-wrap: nowrap !important;
    overflow-x: auto;
}
.st-key-assess_top_nav [data-testid="stColumn"] {
    min-width: 0 !important;
    flex: 0 0 auto !important;
    width: auto !important;
}
.st-key-assess_top_nav button p {
    white-space: nowrap;
}
/* Robot picker row ("Robot:" plus the Xi/SP/DV5 checkboxes): same
   shrink-to-content trick as the top nav above, so the label and all
   three checkboxes sit close together on the left instead of each
   getting an even (and mostly empty) 1/4 of the row's full width —
   plus a tighter gap between them and vertical centering so the
   "Robot:" label lines up with the checkboxes beside it. */
.st-key-assess_robo_row [data-testid="stHorizontalBlock"] {
    flex-wrap: nowrap !important;
    gap: 0.5rem !important;
    align-items: center !important;
}
.st-key-assess_robo_row [data-testid="stColumn"] {
    min-width: 0 !important;
    flex: 0 0 auto !important;
    width: auto !important;
}
/* IMPORTANT: each st.checkbox's own "Xi"/"SP"/"DV5" label text is ALSO
   rendered through a [data-testid="stMarkdownContainer"] > p — the same
   structure the plain "Robot:" text uses — so an unscoped selector
   matches (and repositions) both. Every rule below is scoped with
   :not(:has([data-testid="stCheckbox"])) to hit only the one column
   that's markdown-but-not-a-checkbox ("Robot:" itself), never the
   checkboxes' own internal labels — that's what made every previous
   attempt at nudging "Robot:" also drag Xi/SP/DV5 along with it. */
.st-key-assess_robo_row [data-testid="stColumn"]:has([data-testid="stMarkdownContainer"]):not(:has([data-testid="stCheckbox"])) {
    display: flex;
    align-items: center;
    height: 2.5rem;
}
.st-key-assess_robo_row [data-testid="stColumn"]:has([data-testid="stMarkdownContainer"]):not(:has([data-testid="stCheckbox"]))
    [data-testid="stMarkdownContainer"] {
    margin-bottom: 0 !important;
}
.st-key-assess_robo_row [data-testid="stColumn"]:has([data-testid="stMarkdownContainer"]):not(:has([data-testid="stCheckbox"]))
    [data-testid="stMarkdownContainer"] p {
    white-space: nowrap;
    margin: 0;
    /* Now that this selector is properly isolated from the checkboxes'
       own labels (see the comment above), this transform only ever
       repositions "Robot:" itself. */
    transform: translateY(11px);
}
/* "On mobile: tap the >> icon..." tip: shrink padding, and match the text
   size to "Daily Preparation" and similar field labels (0.75x --pp-substep-font)
   so it's part of the same live, width-responsive size hierarchy instead
   of a fixed size of its own. It's now the first element on its page, so
   pull it up out of Streamlit's default ~96px block-container top padding
   (reserved so content clears the sticky header) — scoped to this
   container only, so it doesn't touch top spacing on any other page. */
.st-key-mobile_tip {
    margin-top: -36px;
}
.st-key-mobile_tip [data-testid="stAlertContainer"] {
    padding-top: 0.35rem;
    padding-bottom: 0.35rem;
    display: flex;
    align-items: center;
}
/* Flex items default to min-width:auto, which refuses to shrink below the
   text's own unwrapped intrinsic width — with white-space:nowrap forced
   below, that let the icon+text row (and everything measuring against it,
   including mobile_tip()'s own fit script) balloon past the real available
   width instead of being constrained by it. min-width:0 at each flex level
   here lets it actually shrink to fit like normal content. */
.st-key-mobile_tip [data-testid="stAlertContainer"],
.st-key-mobile_tip [data-testid="stAlertContentInfo"],
.st-key-mobile_tip [data-testid="stMarkdownContainer"] {
    min-width: 0;
}
.st-key-mobile_tip [data-testid="stAlertContainer"] p {
    font-size: calc(var(--pp-substep-font, 1.3125rem) * 0.75);
    line-height: 1.3;
    /* mobile_tip()'s script shrinks this further so the label never wraps
       past one line; this is the CSS-only fallback if it can't run. */
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
/* Streamlit's alert internals are several nested flex/block layers deep,
   each defaulting to top/stretch alignment, which left the text sitting
   above center within our shorter box. Force every level to center. */
.st-key-mobile_tip [data-testid="stAlertContainer"] * {
    align-items: center !important;
    align-self: center !important;
}
/* Assessment page field labels (Daily Preparation, Overall Performance Rating,
   Development/Improvement/Feed-Forward) and every label inside the
   Step-Level Ratings expander (Case Complexity, then each procedure
   step): sized relative to that expander's own header (--pp-substep-font,
   itself 0.75x the page's main header), so the whole page scales
   together as one hierarchy with window width. */
.st-key-assess_preparation [data-testid="stWidgetLabel"] p,
.st-key-assess_overall_performance [data-testid="stWidgetLabel"] p,
.st-key-assess_notes [data-testid="stWidgetLabel"] p,
.st-key-step_ratings_expander_resident [data-testid="stWidgetLabel"] p,
.st-key-step_ratings_expander_attending [data-testid="stWidgetLabel"] p {
    font-size: calc(var(--pp-substep-font, 1.3125rem) * 0.75);
}
/* Dropdown value text (Daily Preparation, Overall Performance Rating, and every
   Step-Level Ratings entry incl. Case Complexity): 0.8x its own label's
   size above, i.e. 0.6x --pp-substep-font overall. The options popup
   itself renders in a portal under <body>, outside these scoped
   containers, so it isn't reachable here and keeps its default size. */
.st-key-assess_preparation [data-testid="stSelectbox"] input,
.st-key-assess_overall_performance [data-testid="stSelectbox"] input,
.st-key-step_ratings_expander_resident [data-testid="stSelectbox"] input,
.st-key-step_ratings_expander_attending [data-testid="stSelectbox"] input {
    font-size: calc(var(--pp-substep-font, 1.3125rem) * 0.6);
}
/* "In order to improve this:" / "Do this:" — two stacked label+textbox
   rows, one st.columns() call each. sync_improve_how_label_width()
   measures both labels' natural (unwrapped) widths and sets
   --pp-improve-label-width to the wider of the two, so both label
   columns below share that same width regardless of which row's label
   is actually longer — the text boxes' left edges line up between the
   two rows, and each row's label sits snug against its own text box
   instead of at a fixed ratio-based column boundary with a leftover
   gap. Falls back to auto (each column sized to its own content, which
   won't line up between rows) if the measurement script can't run. */
.st-key-assess_improve_how [data-testid="stHorizontalBlock"] {
    align-items: center;
    gap: 0.4rem;
}
.st-key-assess_improve_how [data-testid="stColumn"]:has([data-testid="stMarkdownContainer"]) {
    flex: 0 0 var(--pp-improve-label-width, auto) !important;
    width: var(--pp-improve-label-width, auto) !important;
    min-width: 0 !important;
}
.st-key-assess_improve_how [data-testid="stColumn"]:has([data-testid="stTextInput"]) {
    flex: 1 1 0 !important;
    width: auto !important;
    min-width: 100px !important;
}
/* The two rows are separate blocks stacked in the container's own
   vertical flow — tighten Streamlit's default inter-block gap between
   them so they read as one compact two-line field instead of two
   loosely related rows. */
.st-key-assess_improve_how [data-testid="stVerticalBlock"] {
    gap: 0.3rem;
}
/* Streamlit gives stMarkdownContainer a built-in -16px bottom margin
   (its own vertical-rhythm spacing trick). That negative margin
   collapses this column's layout height down to ~10px even though the
   text still renders at its full ~26px, so the text visually
   overflowed out the bottom of its (collapsed) box — which is what
   made it look bottom-aligned against the input next to it. Cancel it
   so the column's box actually matches its text, and the row's
   align-items:center above can center it correctly. */
.st-key-assess_improve_how [data-testid="stMarkdownContainer"] {
    margin-bottom: 0 !important;
}
.st-key-assess_improve_how [data-testid="stColumn"]:has([data-testid="stMarkdownContainer"]) [data-testid="stMarkdownContainer"] p {
    margin: 0;
    text-align: right;
    font-size: calc(var(--pp-substep-font, 1.3125rem) * 0.75);
}
/* Both rows are separate st.columns() calls, so each is the lone/first
   [data-testid="stHorizontalBlock"] under its own private wrapper —
   :first-of-type can't tell them apart (it matched both, which is why
   that approach briefly left "Do this:" flush left too). The "In order
   to improve this:" label instead gets an inline style="text-align:left"
   straight from Python, which — having higher specificity than this
   class-based rule without needing !important — cleanly overrides just
   that one label while "Do this:" keeps the rule's default right-align. */
.st-key-assess_improve_how [data-testid="stTextInput"] input {
    font-size: calc(var(--pp-substep-font, 1.3125rem) * 0.6);
}
/* The two dividers around this section still carry Streamlit's normal
   ~32px/48px vertical rhythm above/below (each element's own default
   spacing plus the page's inter-element gap) — pull the section's
   outer wrapper up/down with negative margins, unevenly (-16/-32
   rather than a flat -24/-24), so the section actually sits centered
   between the two lines (equal ~16px gaps to each) instead of just
   snug with mismatched gaps. */
[data-testid="stLayoutWrapper"]:has(> .st-key-assess_improve_how) {
    margin-top: -16px !important;
    margin-bottom: -32px !important;
}
/* Same idea for the Overall/Daily Preparation row (Case Complexity now
   lives inside the Step-Level Ratings expander below): its default
   gaps to the divider above (32px) and the expander below (16px)
   weren't equal — pull/push its wrapper so the gap above lands at
   24px and the gap below (bumped an extra 8px per feedback that it
   still looked tighter) at 32px. */
[data-testid="stLayoutWrapper"]:has(> .st-key-assess_ratings_row) {
    margin-top: -8px !important;
    margin-bottom: 16px !important;
}
/* Same idea for the Robot picker row: pull its wrapper down to close
   up most of its default ~32px gap to the divider right below it. */
[data-testid="stLayoutWrapper"]:has(> .st-key-assess_robo_row) {
    margin-bottom: -16px !important;
}
/* Step-Level Ratings expander: a divider now sits between it and
   Development/Improvement/Feed-Forward below, so its default ~48px
   gap to that divider needs pulling in — to 32px, matching the
   divider's own default gap to the field below it, for a symmetric,
   deliberate-looking gap on both sides of the line instead of the
   default's uneven ~48px/32px split. The key class lands directly on
   this expander's own wrapper (unlike a plain st.container), so no
   :has() indirection is needed here. */
.st-key-step_ratings_expander_resident,
.st-key-step_ratings_expander_attending {
    margin-bottom: -16px !important;
}
/* Start page: "Assess Together", "Self-Assessment", and "Blank
   Magic-Link for Attending" all get the same white background /
   bold red border look, overriding whichever primary/secondary
   button type each one otherwise renders as. Label text stays
   plain black, not red/bold, so only the border carries the color. */
.st-key-start_together_btn button,
.st-key-start_self_btn button,
.st-key-start_blank_link_btn button {
    background-color: #FFFFFF !important;
    border: 3px solid #FF4B4B !important;
    color: #000000 !important;
}
.st-key-start_together_btn button p,
.st-key-start_self_btn button p,
.st-key-start_blank_link_btn button p {
    color: #000000 !important;
    font-weight: normal !important;
}
.st-key-start_together_btn button:hover,
.st-key-start_self_btn button:hover,
.st-key-start_blank_link_btn button:hover {
    background-color: #FFF0F0 !important;
    border-color: #E63946 !important;
    color: #000000 !important;
}
</style>
""",
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────
# PAGE ROUTER
# ─────────────────────────────────────────────
page = st.session_state["page"]


# ════════════════════════════════════════════════════════════
# PAGE: LOGIN
# ════════════════════════════════════════════════════════════
if page == "login":

    def _complete_login(canonical_email: str) -> None:
        """Password verified (or just created) — resolve the account and
        drop into the app proper."""
        residents = read_sheet_df(
            SHEET_RESIDENTS, expected_cols=["email", "name", "specialty_id", "created_at"]
        )
        admins_lower = [a.lower() for a in ADMINS]
        email_lower = canonical_email.strip().lower()
        residents_lower = residents["email"].str.strip().str.lower()
        if email_lower in admins_lower:
            st.session_state.update(resident=canonical_email, resident_name="Admin", role="admin", page="admin")
        elif email_lower in residents_lower.values:
            row = residents.loc[residents_lower == email_lower].iloc[0]
            st.session_state.update(
                resident=row["email"], resident_name=row["name"],
                specialty_id=row["specialty_id"], role="resident", page="home",
            )
        else:
            attendings = _read_attendings_df()
            attendings_lower = attendings["email"].fillna("").str.strip().str.lower()
            row = attendings.loc[attendings_lower == email_lower].iloc[0]
            st.session_state.update(
                role="attending",
                attending_login_email=row["email"],
                attending_login_name=row["attending_name"],
                attending_login_id=row["attending_id"],
                attending_login_specialty_id=row["specialty_id"],
                page="attending_home",
            )
        st.rerun()

    page_header("🩺 Procedure Passport")
    st.markdown("_Track your surgical skills journey, one procedure at a time._")
    st.markdown("---")

    # Email and password on one screen, submitted together — whether
    # this is a first-ever login (no password on file yet) or a
    # returning one is only known once the email is looked up, which
    # only happens on submit, so both fields are always shown up
    # front rather than password appearing as a separate step after
    # the email is entered.
    email = st.text_input("Email address", placeholder="you@hospital.org", key="login_email_input")
    pw = st.text_input("Password", type="password", key="login_pw_input")
    st.caption(
        "First time here, or no password set yet? Leave Password blank and "
        "click Log In — you'll be prompted to choose one (8+ characters)."
    )
    if st.button("Log In →", width="stretch", type="primary"):
        if not email.strip():
            st.error("Please enter your email address.")
        else:
            # Password isn't required just to submit — whether one's
            # even needed depends on whether this account has one on
            # file yet, which isn't known until after the lookup
            # below. A first-time user can submit with the password
            # field left blank; they're told to choose one once
            # that's confirmed, without ever leaving this page.
            try:
                residents = read_sheet_df(
                    SHEET_RESIDENTS,
                    expected_cols=["email", "name", "specialty_id", "created_at"],
                )
                attendings = _read_attendings_df()
                email_lower = email.strip().lower()
                admins_lower = [a.lower() for a in ADMINS]
                residents_lower = residents["email"].str.strip().str.lower()
                attendings_lower = attendings["email"].fillna("").str.strip().str.lower()
                if email_lower in admins_lower:
                    canonical = ADMINS[admins_lower.index(email_lower)]
                elif email_lower in residents_lower.values:
                    canonical = residents.loc[residents_lower == email_lower].iloc[0]["email"]
                elif email_lower in attendings_lower.values:
                    canonical = attendings.loc[attendings_lower == email_lower].iloc[0]["email"]
                else:
                    canonical = None
                if canonical is None:
                    st.error("❌ Email not recognised. Ask an admin to add you.")
                elif get_password_row(canonical) is not None:
                    if not pw:
                        st.error("Please enter your password.")
                    elif verify_password(canonical, pw):
                        _complete_login(canonical)
                    else:
                        st.error("❌ Incorrect password.")
                elif len(pw) < 8:
                    st.info(
                        f"👋 First time logging in as **{canonical}** — "
                        "enter a password above (8+ characters) and log in again "
                        "to set it as your password."
                    )
                else:
                    set_password(canonical, pw)
                    _complete_login(canonical)
            except ConnectionError as exc:
                show_gs_error(exc)


# ════════════════════════════════════════════════════════════
# PAGE: ADMIN PANEL
# ════════════════════════════════════════════════════════════
elif page == "admin":
    page_header("⚙️ Admin Panel")
    if st.button("🏠 Back to Home", key="admin_home_top"):
        go_to("home")

    if st.button("🔄 Reload Data"):
        st.cache_data.clear()
        st.rerun()

    # ── Specialties ──────────────────────────────────────
    st.subheader("Specialties")
    try:
        specialties = read_sheet_df(SHEET_SPECIALTY, expected_cols=["specialty_id", "specialty_name"])
        st.dataframe(specialties, width="stretch")

        with st.expander("➕ Add Specialty"):
            new_spec_id   = st.text_input("Specialty ID (e.g., GS)")
            new_spec_name = st.text_input("Specialty name (e.g., General Surgery)")
            if st.button("Add Specialty", key="btn_add_spec"):
                if new_spec_id and new_spec_name:
                    if new_spec_id in specialties["specialty_id"].values:
                        st.warning("That ID already exists.")
                    else:
                        specialties = pd.concat(
                            [specialties, pd.DataFrame([{"specialty_id": new_spec_id,
                                                          "specialty_name": new_spec_name}])],
                            ignore_index=True,
                        )
                        write_sheet_df(SHEET_SPECIALTY, specialties)
                        st.success(f"✅ Added {new_spec_name}")
                        time.sleep(0.5)
                        st.rerun()
                else:
                    st.error("Please fill in both fields.")
    except ConnectionError as exc:
        show_gs_error(exc)

    st.markdown("---")

    # ── Residents ────────────────────────────────────────
    st.subheader("Residents")
    try:
        spec_df = read_sheet_df(SHEET_SPECIALTY, expected_cols=["specialty_id", "specialty_name"])
        spec_name_to_id = dict(zip(spec_df["specialty_name"], spec_df["specialty_id"]))

        residents = read_sheet_df(
            SHEET_RESIDENTS, expected_cols=["email", "name", "specialty_id", "created_at"]
        )
        disp = residents.merge(spec_df, how="left", on="specialty_id")
        st.dataframe(disp[["email", "name", "specialty_name", "created_at"]], width="stretch")

        with st.expander("➕ Add Resident"):
            new_res_email = st.text_input("Email")
            new_res_name  = st.text_input("Full name")
            new_res_spec  = st.selectbox("Specialty", list(spec_name_to_id.keys()), key="add_res_spec")
            if st.button("Add Resident", key="btn_add_res"):
                if new_res_email and new_res_name and new_res_spec:
                    ensure_resident(new_res_email, new_res_name, spec_name_to_id[new_res_spec])
                    st.success(f"✅ Added {new_res_email}")
                    time.sleep(0.5)
                    st.rerun()
                else:
                    st.warning("Please fill in all fields.")

        if not residents.empty:
            with st.expander("🔑 Reset Password"):
                st.caption("Clears their stored password — their next login will prompt them to set a new one.")
                reset_email = st.selectbox("Select resident", residents["email"], key="reset_res_pw")
                if st.button("Reset Password", key="btn_reset_res_pw"):
                    clear_password(reset_email)
                    st.success(f"✅ Password cleared for {reset_email}")
                    time.sleep(0.5)
                    st.rerun()

            with st.expander("🗑️ Delete Resident"):
                del_email = st.selectbox("Select resident to delete", residents["email"], key="del_res")
                if st.button("Delete", key="btn_del_res"):
                    updated = residents[residents["email"] != del_email].reset_index(drop=True)
                    write_sheet_df(SHEET_RESIDENTS, updated)
                    clear_password(del_email)
                    st.success(f"Deleted {del_email}")
                    time.sleep(0.5)
                    st.rerun()
    except ConnectionError as exc:
        show_gs_error(exc)

    st.markdown("---")

    # ── My Account ─────────────────────────────────────────
    st.subheader("My Account")
    with st.expander("🔑 Reset My Password"):
        st.caption(
            "Clears your own stored password — you'll set a new one the next time you log in. "
            "You'll stay logged in for this session."
        )
        if st.button("Reset My Password", key="btn_reset_own_pw"):
            try:
                clear_password(st.session_state["resident"])
                st.success("✅ Your password has been cleared. You'll set a new one next time you log in.")
            except ConnectionError as exc:
                show_gs_error(exc)

    st.markdown("---")

    # ── Attendings ───────────────────────────────────────
    st.subheader("Attendings")
    try:
        attendings = _read_attendings_df()
        spec_df, _, _, _ = load_refs()
        st.dataframe(attendings, width="stretch")

        with st.expander("➕ Add Attending"):
            new_att_name  = st.text_input("Attending name")
            new_att_spec  = st.selectbox("Specialty", spec_df["specialty_name"], key="add_att_spec")
            new_att_email = st.text_input(
                "Email (optional)",
                help="Set this to give the attending their own login — they'll see only "
                     "the residents in their specialty, and can start a blank assessment "
                     "or view a resident's dashboard directly (no magic link needed).",
            )
            if st.button("Add Attending", key="btn_add_att"):
                if new_att_name:
                    _spec_match = spec_df[spec_df["specialty_name"].astype(str).str.strip() == str(new_att_spec).strip()]
                    spec_id = _spec_match["specialty_id"].values[0] if len(_spec_match) > 0 else None
                    ensure_attending(new_att_name, spec_id, new_att_email)
                    st.success(f"✅ Added {new_att_name}")
                    time.sleep(0.5)
                    st.rerun()
                else:
                    st.error("Please enter an attending name.")

        if not attendings.empty:
            with st.expander("✏️ Edit Attending / Assign Login"):
                st.caption(
                    "Give an existing attending an email here to designate their account "
                    "as an attending login — they'll then be able to log in themselves "
                    "(same Log In screen as residents) and get their own Attending Home, "
                    "Start Assessment, and Resident Dashboard."
                )
                edit_att_name = st.selectbox(
                    "Select attending to edit", attendings["attending_name"], key="edit_att_select"
                )
                _edit_row = attendings[attendings["attending_name"] == edit_att_name].iloc[0]
                _edit_spec_name_match = spec_df.loc[
                    spec_df["specialty_id"] == _edit_row["specialty_id"], "specialty_name"
                ]
                _edit_spec_default = _edit_spec_name_match.values[0] if len(_edit_spec_name_match) else None
                _edit_spec_options = list(spec_df["specialty_name"])
                _edit_spec_idx = (
                    _edit_spec_options.index(_edit_spec_default)
                    if _edit_spec_default in _edit_spec_options else 0
                )
                edit_att_name_new = st.text_input(
                    "Attending name", value=str(_edit_row["attending_name"]), key="edit_att_name"
                )
                edit_att_spec = st.selectbox(
                    "Specialty", _edit_spec_options, index=_edit_spec_idx, key="edit_att_spec"
                )
                edit_att_email = st.text_input(
                    "Email (blank = no login)",
                    value="" if pd.isna(_edit_row["email"]) else str(_edit_row["email"]),
                    key="edit_att_email",
                )
                if st.button("Save Changes", key="btn_edit_att"):
                    if not edit_att_name_new.strip():
                        st.error("Please enter an attending name.")
                    else:
                        _spec_match = spec_df[spec_df["specialty_name"].astype(str).str.strip() == str(edit_att_spec).strip()]
                        _new_spec_id = _spec_match["specialty_id"].values[0] if len(_spec_match) > 0 else None
                        _old_email = "" if pd.isna(_edit_row["email"]) else str(_edit_row["email"]).strip()
                        _new_email = edit_att_email.strip()
                        updated = attendings.copy()
                        _mask = updated["attending_name"] == edit_att_name
                        updated.loc[_mask, "attending_name"] = edit_att_name_new.strip()
                        updated.loc[_mask, "specialty_id"]   = _new_spec_id
                        updated.loc[_mask, "email"]          = _new_email
                        write_sheet_df(SHEET_ATTENDINGS, updated)
                        # Changing/removing the email invalidates any password stored
                        # under the old one — it would otherwise be an orphaned,
                        # unreachable credential.
                        if _old_email and _old_email.lower() != _new_email.lower():
                            clear_password(_old_email)
                        st.success(f"✅ Saved {edit_att_name_new.strip()}")
                        time.sleep(0.5)
                        st.rerun()

            with st.expander("🔑 Reset Attending Password"):
                st.caption("Clears their stored password — their next login will prompt them to set a new one.")
                _att_with_email = attendings[attendings["email"].fillna("").astype(str).str.strip() != ""]
                if _att_with_email.empty:
                    st.caption("_No attendings have a login email set yet — add one under Edit Attending above._")
                else:
                    reset_att = st.selectbox(
                        "Select attending", _att_with_email["attending_name"], key="reset_att_pw"
                    )
                    if st.button("Reset Password", key="btn_reset_att_pw"):
                        _reset_email = _att_with_email.loc[
                            _att_with_email["attending_name"] == reset_att, "email"
                        ].values[0]
                        clear_password(str(_reset_email))
                        st.success(f"✅ Password cleared for {reset_att}")
                        time.sleep(0.5)
                        st.rerun()

            with st.expander("🗑️ Delete Attending"):
                del_att = st.selectbox("Select attending to delete", attendings["attending_name"], key="del_att")
                if st.button("Delete", key="btn_del_att"):
                    _del_row = attendings[attendings["attending_name"] == del_att].iloc[0]
                    _del_email = "" if pd.isna(_del_row["email"]) else str(_del_row["email"]).strip()
                    updated = attendings[attendings["attending_name"] != del_att].reset_index(drop=True)
                    write_sheet_df(SHEET_ATTENDINGS, updated)
                    if _del_email:
                        clear_password(_del_email)
                    st.success(f"Deleted {del_att}")
                    time.sleep(0.5)
                    st.rerun()
    except ConnectionError as exc:
        show_gs_error(exc)

    st.markdown("---")

    # ── Procedures ───────────────────────────────────────
    st.subheader("Procedures")
    try:
        spec_df, _, _, _ = load_refs()

        with st.expander("➕ Add New Procedure"):
            new_proc_id   = st.text_input("Procedure ID (e.g., CSEC)").strip().upper()
            new_proc_name = st.text_input("Procedure name (e.g., Cesarean Section)")
            new_proc_spec = st.selectbox("Specialty", spec_df["specialty_name"], key="add_proc_spec")
            steps_raw     = st.text_area("Steps (one per line)")
            new_steps     = [s.strip() for s in steps_raw.split("\n") if s.strip()]
            if st.button("Add Procedure", key="btn_add_proc"):
                if new_proc_id and new_proc_name and new_steps:
                    _spec_match = spec_df[spec_df["specialty_name"].astype(str).str.strip() == str(new_proc_spec).strip()]
                    spec_id = _spec_match["specialty_id"].values[0] if len(_spec_match) > 0 else None
                    ensure_procedure(new_proc_id, new_proc_name, spec_id, new_steps)
                    st.success(f"✅ Added {new_proc_name}")
                    time.sleep(0.5)
                    st.rerun()
                else:
                    st.error("Please fill in all fields and at least one step.")

        with st.expander("✏️ Edit Existing Procedure"):
            procs_df = read_sheet_df(SHEET_PROCEDURES, expected_cols=["procedure_id", "procedure_name", "specialty_id"])
            if procs_df.empty:
                st.info("No procedures yet.")
            else:
                edit_proc    = st.selectbox("Select procedure", procs_df["procedure_name"], key="edit_proc_sel")
                _proc_match = procs_df[procs_df["procedure_name"].astype(str).str.strip() == str(edit_proc).strip()]
                sel_proc_id  = _proc_match["procedure_id"].values[0] if len(_proc_match) > 0 else None
                new_pname    = st.text_input("Updated name", value=edit_proc, key="edit_proc_name")

                # Pre-filled, per-step editor rather than retyping the whole
                # step list as plain text: renaming or reordering a step no
                # longer regenerates its step_id (which used to bake in its
                # position, e.g. S_CIRC_01/02/03 — so ANY edit anywhere in
                # the list shifted every step after it onto a new ID, silently
                # cutting that step's historical ratings loose from it). Order
                # is a plain editable number rather than relying on drag-to-
                # reorder (data_editor doesn't support that) — retype a row's
                # Order to move it anywhere, including fractional values like
                # 2.5 to insert a new step between steps 2 and 3.
                _all_steps_df   = read_sheet_df(
                    SHEET_STEPS, expected_cols=["step_id", "procedure_id", "step_order", "step_name"]
                )
                _proc_steps_df  = _all_steps_df[_all_steps_df["procedure_id"] == sel_proc_id].sort_values("step_order")
                _existing_step_ids = set(_proc_steps_df["step_id"])
                _steps_editor_df = pd.DataFrame({
                    "step_id": _proc_steps_df["step_id"].tolist(),
                    "Order":   list(range(1, len(_proc_steps_df) + 1)),
                    "Step":    _proc_steps_df["step_name"].tolist(),
                })
                st.caption(
                    "Rename or reorder existing steps — their rating history stays "
                    "linked either way. Add a row for a new step; give it an Order "
                    "number to place it anywhere (e.g. 2.5 inserts it between "
                    "steps 2 and 3). Delete a row to remove that step."
                )
                _edited_steps_df = st.data_editor(
                    _steps_editor_df,
                    column_config={
                        "step_id": None,  # identity only — never shown or hand-edited
                        "Order": st.column_config.NumberColumn(
                            "Order", help="Position in the sequence. Fractional values are fine.",
                        ),
                        "Step": st.column_config.TextColumn("Step", required=True),
                    },
                    num_rows="dynamic",
                    hide_index=True,
                    width="stretch",
                    key=f"edit_proc_steps_editor_{sel_proc_id}",
                )

                if st.button("Update Procedure", key="btn_upd_proc"):
                    procs_df.loc[procs_df["procedure_id"] == sel_proc_id, "procedure_name"] = new_pname
                    write_sheet_df(SHEET_PROCEDURES, procs_df)

                    _clean_steps = _edited_steps_df.dropna(subset=["Step"]).copy()
                    _clean_steps = _clean_steps[_clean_steps["Step"].astype(str).str.strip() != ""]
                    if _clean_steps.empty:
                        st.error("A procedure needs at least one step — add one before updating.")
                    else:
                        # Stable sort: rows with no Order (a just-added row
                        # left blank) fall to the end rather than the top.
                        _clean_steps["Order"] = pd.to_numeric(_clean_steps["Order"], errors="coerce")
                        _blank_order_fill = (
                            _clean_steps["Order"].max() + 1 if _clean_steps["Order"].notna().any() else 1
                        )
                        _clean_steps["Order"] = _clean_steps["Order"].fillna(_blank_order_fill)
                        _clean_steps = _clean_steps.sort_values("Order", kind="stable").reset_index(drop=True)

                        _new_step_rows = []
                        for i, _row in _clean_steps.iterrows():
                            _sid = _row["step_id"]
                            if not isinstance(_sid, str) or not _sid.strip() or _sid not in _existing_step_ids:
                                _sid = f"S_{sel_proc_id}_{uuid.uuid4().hex[:8]}"
                            _new_step_rows.append({
                                "step_id":      _sid,
                                "procedure_id": sel_proc_id,
                                "step_order":   i + 1,
                                "step_name":    str(_row["Step"]).strip(),
                            })
                        updated_steps = pd.DataFrame(_new_step_rows)
                        steps_df = _all_steps_df[_all_steps_df["procedure_id"] != sel_proc_id]
                        steps_df = pd.concat([steps_df, updated_steps], ignore_index=True)
                        write_sheet_df(SHEET_STEPS, steps_df)
                        st.success(f"✅ Updated '{new_pname}'")
                        time.sleep(0.5)
                        st.rerun()

        # Read-only lookup, deliberately separate from the editor above —
        # for spotting a step name (e.g. "Case Preparation") reused, or
        # near-duplicated, across procedures before deciding which ones to
        # go edit. Doesn't touch any data.
        with st.expander("🔍 Find Steps by Name"):
            st.caption(
                "Pick one or more step names — each exact name that appears "
                "anywhere shows every procedure that has it, so you can spot "
                "one reused (or near-duplicated, e.g. \"Case Preparation\" vs. "
                "\"Daily Preparation\") across procedures before deciding which "
                "to go edit or delete."
            )
            _search_steps_df = read_sheet_df(
                SHEET_STEPS, expected_cols=["step_id", "procedure_id", "step_order", "step_name"]
            )
            _all_step_names = sorted(
                _search_steps_df["step_name"].dropna().astype(str).str.strip().unique().tolist(),
                key=str.lower,
            )
            _search_terms = st.multiselect(
                "Step name(s) to search", _all_step_names, key="step_search_terms",
            )
            if _search_terms:
                _proc_name_lookup = dict(zip(procs_df["procedure_id"], procs_df["procedure_name"]))
                for term in _search_terms:
                    _hits = _search_steps_df[
                        _search_steps_df["step_name"].astype(str).str.strip() == term
                    ].sort_values(["procedure_id", "step_order"])
                    st.markdown(f"**\"{term}\"** — {len(_hits)} match{'es' if len(_hits) != 1 else ''}")
                    _hits_display = pd.DataFrame({
                        "Procedure": _hits["procedure_id"].map(_proc_name_lookup).fillna(_hits["procedure_id"]),
                        "Step":      _hits["step_name"],
                        "Order":     _hits["step_order"],
                    })
                    st.dataframe(_hits_display, width="stretch", hide_index=True)
    except ConnectionError as exc:
        show_gs_error(exc)

    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("⬅️ Back to Login"):
            go_to("login")
    with col2:
        if st.button("🏠 Resident Home"):
            go_to("home")


# ════════════════════════════════════════════════════════════
# PAGE: HOME
# ════════════════════════════════════════════════════════════
elif page == "home":
    mobile_tip("📱 On mobile: tap the >> icon at top left to access navigation and rating legend.")
    # tier_text excludes the resident's name from _header_max()'s length
    # tier so a long name doesn't needlessly drop the whole header into a
    # smaller ceiling; header_break_before() also protects every other
    # space so that if this does wrap, it can only break right after the
    # comma, never mid-name.
    page_header(
        header_break_before("👋 Welcome back,", st.session_state["resident_name"]),
        tier_text="👋 Welcome back,",
    )
    st.markdown("_What would you like to do today?_")
    st.markdown("")

    with st.container(key="home_cards"):
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown('<div class="pp-card">', unsafe_allow_html=True)
            st.markdown("### ➕ New Assessment")
            st.markdown("Start a new procedure case and record step ratings.")
            if st.button("Start Assessment", width="stretch", type="primary"):
                go_to("start")
            st.markdown("</div>", unsafe_allow_html=True)

        with c2:
            st.markdown('<div class="pp-card">', unsafe_allow_html=True)
            st.markdown("### 📊 Cumulative Dashboard")
            st.markdown("View your progress heatmap over time.")
            if st.button("View Dashboard", width="stretch"):
                go_to("cumulative")
            st.markdown("</div>", unsafe_allow_html=True)

        with c3:
            st.markdown('<div class="pp-card">', unsafe_allow_html=True)
            st.markdown("### 💬 Comments")
            st.markdown("Browse and export all attending feedback.")
            if st.button("View Comments", width="stretch"):
                go_to("comments")
            st.markdown("</div>", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
# PAGE: ATTENDING HOME
# ════════════════════════════════════════════════════════════
elif page == "attending_home":
    mobile_tip("📱 On mobile: tap the >> icon at top left to access navigation.")
    page_header(
        header_break_before("👋 Welcome back,", st.session_state["attending_login_name"]),
        tier_text="👋 Welcome back,",
    )
    st.markdown("_What would you like to do today?_")
    st.markdown("")

    with st.container(key="home_cards"):
        c1, c2 = st.columns(2)
        with c1:
            st.markdown('<div class="pp-card">', unsafe_allow_html=True)
            st.markdown("### ➕ New Assessment")
            st.markdown("Start a blank assessment for one of your residents.")
            if st.button("Start Assessment", width="stretch", type="primary", key="att_home_start_btn"):
                go_to("attending_start")
            st.markdown("</div>", unsafe_allow_html=True)

        with c2:
            st.markdown('<div class="pp-card">', unsafe_allow_html=True)
            st.markdown("### 📊 Resident Dashboard")
            st.markdown("View resident progress heatmaps and comments.")
            if st.button("View Dashboard", width="stretch", key="att_home_dashboard_btn"):
                go_to("attending_resident_dashboard")
            st.markdown("</div>", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
# PAGE: START CASE
# ════════════════════════════════════════════════════════════
elif page == "start":
    mobile_tip("📱 On mobile: tap the >> icon at top left to view the sidebar.")
    page_header("📋 Start Assessment")
    if st.button("🏠 Back to Home", key="start_home_top"):
        go_to("home")

    try:
        spec_df, proc_df, steps_df, atnd_df = load_refs()
    except ConnectionError as exc:
        show_gs_error(exc)
        if st.button("⬅️ Back to Home"):
            go_to("home")
        st.stop()

    spec_map = dict(zip(spec_df["specialty_name"], spec_df["specialty_id"]))
    is_admin = st.session_state["resident"] in ADMINS

    if is_admin:
        selected_spec_name = st.selectbox("Specialty", list(spec_map.keys()))
        specialty_id       = spec_map[selected_spec_name]
        st.session_state["specialty_id"] = specialty_id
    else:
        specialty_id = st.session_state.get("specialty_id")
        if specialty_id is None:
            st.error("No specialty assigned. Contact an admin.")
            st.stop()

    procs = proc_df[proc_df["specialty_id"] == specialty_id]
    atnds = atnd_df[atnd_df["specialty_id"] == specialty_id]

    if procs.empty:
        st.warning("⚠️ No procedures configured for this specialty.")
        if st.button("⬅️ Back to Home"):
            go_to("home")
        st.stop()
    if atnds.empty:
        st.warning("⚠️ No attendings configured for this specialty.")
        if st.button("⬅️ Back to Home"):
            go_to("home")
        st.stop()

    proc_map = dict(zip(procs["procedure_name"], procs["procedure_id"]))
    atnd_map = dict(zip(atnds["attending_name"], atnds["attending_id"]))

    _CHOOSE_PROC = "Choose Procedure"
    _CHOOSE_ATT  = "Choose Attending"

    _proc_options = _ordered_procedure_names(proc_map)

    attending = st.selectbox(
        "Attending",
        [_CHOOSE_ATT] + sorted(atnd_map.keys(), key=lambda n: n.split()[-1] if n.split() else n),
    )
    procedure = st.selectbox("Procedure", [_CHOOSE_PROC] + _proc_options)
    case_date = st.date_input("Date", st.session_state["date"])

    procedure_chosen = procedure != _CHOOSE_PROC
    attending_chosen = attending != _CHOOSE_ATT

    st.session_state["procedure_id"] = proc_map[procedure] if procedure_chosen else None
    st.session_state["attending_id"] = atnd_map[attending] if attending_chosen else None
    st.session_state["date"]         = case_date

    if not (procedure_chosen and attending_chosen):
        st.info("Choose an attending and a procedure to continue.")

    st.markdown("---")

    def _reset_and_start(mode: str):
        st.session_state["scores"]               = {}
        st.session_state["notes"]                = ""
        st.session_state["improve"]              = ""
        st.session_state["how"]                  = ""
        st.session_state["generated_magic_link"] = None
        st.session_state["assessment_mode"]      = mode
        go_to("assessment")

    _selection_incomplete = not (procedure_chosen and attending_chosen)

    # Self-Assess's label is the longest of the three, so give it more of
    # the row's width and take it from the other two — fit_all_button_labels()
    # still shrinks per-button as a backstop, but this keeps all three
    # legible at typical widths instead of relying on that shrink alone.
    _start_cols = st.columns([1] if is_admin else [0.85, 1.3, 0.85])
    with _start_cols[0]:
        # Two trailing spaces before \n is Markdown's hard-break syntax —
        # renders as a real <br>, forcing exactly two lines regardless of
        # width (never collapses to one, and white-space:nowrap on
        # button p keeps either line from wrapping into a third).
        # fit_all_button_labels() shrinks the font if the wider of the
        # two lines would otherwise overflow, so neither line ever ends
        # in an ellipsis either — verified at both desktop and cramped
        # widths before applying. Every emoji sits at the end of its
        # own line's text (not before, not centered separately from
        # it) — simpler and it just reads as part of that line.
        if st.button("Assess Together ✅  \nResident + Attending", type="primary", width="stretch", key="start_together_btn", disabled=_selection_incomplete):
            _reset_and_start("together")

    if not is_admin:
        with _start_cols[1]:
            if st.button("Self-Assess ✅  \nPre-Filled Magic Link for Attending 🔗", width="stretch", key="start_self_btn", disabled=_selection_incomplete):
                _reset_and_start("self")
        with _start_cols[2]:
            if st.button("Blank Magic Link 🔗  \nSend to Attending ✉️", width="stretch", key="start_blank_link_btn", disabled=_selection_incomplete):
                _att_match = atnds[atnds["attending_id"].astype(str).str.strip()
                                    == str(st.session_state.get("attending_id", "")).strip()]
                safe_att = _att_match["attending_name"].values[0].replace(" ", "_") if len(_att_match) > 0 else "Unknown"
                base_url = st.secrets.get("APP_BASE_URL", "https://procedurepassport.streamlit.app")
                st.session_state["blank_magic_link"] = (
                    f"{base_url}/?mode=attending"
                    f"&resident={st.session_state['resident']}"
                    f"&procedure_id={st.session_state['procedure_id']}"
                    f"&specialty_id={specialty_id}"
                    f"&attending_name={safe_att}"
                    f"&date={st.session_state['date']}"
                )

    if st.session_state.get("blank_magic_link"):
        st.success("✅ A blank link is ready for your attending:")
        copy_link_button(st.session_state["blank_magic_link"], key="copy_blank_link")
        st.code(st.session_state["blank_magic_link"], language="text")

    st.markdown("---")
    if st.button("⬅️ Back to Home"):
        go_to("home")


# ════════════════════════════════════════════════════════════
# PAGE: ASSESSMENT
# ════════════════════════════════════════════════════════════
elif page == "assessment":
    try:
        _, proc_df, steps_df, atnd_df = load_refs()
    except ConnectionError as exc:
        show_gs_error(exc)
        if st.button("⬅️ Back to Start"):
            go_to("start")
        st.stop()

    is_admin = st.session_state["resident"] in ADMINS

    steps = steps_df[steps_df["procedure_id"] == st.session_state["procedure_id"]].sort_values("step_order")
    if steps.empty:
        st.error("No steps defined for this procedure. Ask an admin to add steps.")
        if st.button("⬅️ Back to Start"):
            go_to("start")
        st.stop()

    # Resolve procedure name for the page title (Fix 3)
    _proc_rows = proc_df.loc[proc_df["procedure_id"] == st.session_state["procedure_id"], "procedure_name"].values
    _proc_name = _proc_rows[0] if len(_proc_rows) else "Assessment"
    mobile_tip("📱 On mobile: tap the >> icon at top left to view the sidebar.")
    # tier_text excludes the resident's name from _header_max()'s length
    # tier so a long name doesn't needlessly drop the header into a
    # smaller ceiling; the fit script still measures and shrinks the
    # full displayed text (name included) if it doesn't actually fit.
    page_header(
        header_break_before(f"📝 {_proc_name}", f"Assessment for {st.session_state['resident_name']}"),
        tier_text=f"📝 {_proc_name} Assessment",
    )
    assessment_instructions_note()

    # Back button placed at the top, clearly separated from Finish (Fix 7)
    with st.container(key="assess_top_nav"):
        _top_cols_assess = st.columns([1, 1, 4])
        with _top_cols_assess[0]:
            if st.button("⬅️ Back to Start", key="back_top"):
                go_to("start")
        with _top_cols_assess[1]:
            if st.button("🏠 Home", key="assess_home_top"):
                go_to("home")

    if _is_robotic_procedure(_proc_name):
        render_robo_type_picker("robo_type", default=st.session_state.get("robo_type", "Xi"))

    st.markdown("---")

    with st.container(key="assess_improve_how"):
        _imp_label_col, _imp_input_col = st.columns([2, 6])
        with _imp_label_col:
            st.markdown(
                '<p style="text-align: left;">In order to improve this:</p>',
                unsafe_allow_html=True,
            )
        with _imp_input_col:
            st.session_state["improve"] = st.text_input(
                "What to improve",
                value=st.session_state.get("improve", ""),
                key="assess_improve",
                label_visibility="collapsed",
                placeholder="e.g., suture technique",
            )
        _how_label_col, _how_input_col = st.columns([2, 6])
        with _how_label_col:
            st.markdown("Do this:")
        with _how_input_col:
            st.session_state["how"] = st.text_input(
                "How to improve it",
                value=st.session_state.get("how", ""),
                key="assess_how",
                label_visibility="collapsed",
                placeholder="e.g., practice two-handed knots",
            )
    sync_improve_how_label_width()

    st.markdown("---")

    with st.container(key="assess_ratings_row"):
        _overall_col, _prep_col = st.columns(2)
        with _overall_col:
            current_o = st.session_state.get("overall_performance", O_SCORE_OPTIONS[0])
            st.session_state["overall_performance"] = st.selectbox(
                "Overall Performance Rating",
                O_SCORE_OPTIONS,
                index=O_SCORE_OPTIONS.index(current_o) if current_o in O_SCORE_OPTIONS else 0,
                key="assess_overall_performance",
            )
        with _prep_col:
            _cp_opts = ["Not Assessed", "Unprepared", "Poorly Prepared",
                        "Adequately Prepared", "Well Prepared", "Highly Prepared"]
            _cp_default = st.session_state.get("case_preparation", "Not Assessed")
            _cp_idx = _cp_opts.index(_cp_default) if _cp_default in _cp_opts else 0
            st.session_state["case_preparation"] = st.selectbox(
                "Daily Preparation",
                _cp_opts,
                index=_cp_idx,
                key="assess_preparation",
            )

    with st.expander(
        header_break_before("Step-Level Ratings for", _proc_name),
        expanded=False,
        key="step_ratings_expander_resident",
    ):
        # Case Complexity leads the Step-Level Ratings section, then each
        # procedure step in order.
        _cc_opts = ["— Select complexity —", "Straight Forward", "Moderate", "Complex"]
        _cc_default = st.session_state.get("case_complexity", "— Select complexity —")
        _cc_idx = _cc_opts.index(_cc_default) if _cc_default in _cc_opts else 0
        st.session_state["case_complexity"] = st.selectbox(
            "Case Complexity",
            _cc_opts,
            index=_cc_idx,
            key="assess_case_complexity",
        )
        # Fix 6: reverting to "Not Assessed" is supported — "Not Assessed" is index 0
        # in RATING_OPTIONS so the user can always select it from the dropdown.
        for _, row in steps.iterrows():
            step_id   = row["step_id"]
            step_name = row["step_name"]
            current   = st.session_state["scores"].get(step_id, "Not Assessed")
            st.session_state["scores"][step_id] = st.selectbox(
                step_name,
                RATING_OPTIONS,
                index=RATING_OPTIONS.index(current) if current in RATING_OPTIONS else 0,
                key=f"score_{step_id}",
            )

    st.markdown("---")

    st.session_state["notes"] = st.text_area(
        "Development / Improvement / Feed-Forward", st.session_state.get("notes", ""), key="assess_notes"
    )

    def _assessment_has_value() -> bool:
        return (
            st.session_state["case_complexity"] != "— Select complexity —"
            or st.session_state["case_preparation"] != "Not Assessed"
            or st.session_state["overall_performance"] != O_SCORE_OPTIONS[0]
            or any(v != "Not Assessed" for v in st.session_state["scores"].values())
            or st.session_state.get("notes", "").strip() != ""
            or st.session_state.get("improve", "").strip() != ""
            or st.session_state.get("how", "").strip() != ""
        )

    # Only meaningful for a robotic procedure — st.session_state["robo_type"]
    # could otherwise still hold a stale Xi/SP/DV5 pick left over from a
    # different, earlier robotic procedure this same session.
    _robo_type_to_save = st.session_state.get("robo_type") if _is_robotic_procedure(_proc_name) else None

    def _save_own_case(assessment_type: str) -> str:
        """Save the resident's own entry, tagged with how it was taken."""
        return save_case(
            resident_email=st.session_state["resident"],
            date=st.session_state["date"],
            specialty_id=st.session_state["specialty_id"],
            procedure_id=st.session_state["procedure_id"],
            attending_id=st.session_state["attending_id"],
            scores_dict=st.session_state["scores"],
            case_complexity=st.session_state["case_complexity"],
            case_preparation=st.session_state["case_preparation"],
            overall_performance=st.session_state["overall_performance"],
            robo_type=_robo_type_to_save,
            notes=st.session_state.get("notes", ""),
            improve=st.session_state.get("improve", ""),
            how=st.session_state.get("how", ""),
            assessment_type=assessment_type,
        )

    # Admins never see the magic-link options (Start page hides those
    # buttons for them), so treat any admin session as "together" too,
    # regardless of whatever assessment_mode happens to be stored.
    _mode = "self" if (not is_admin and st.session_state.get("assessment_mode") == "self") else "together"

    st.markdown("---")

    if _mode == "together":
        # Fix 7: Finish button alone at the bottom with a confirmation note
        st.caption("✅ The case is saved automatically when you click Finish & Save.")
        if st.button("🏁 Finish & Save →", type="primary", width="stretch"):
            if not _assessment_has_value():
                st.warning("Please provide at least one rating or comment before submitting.")
            else:
                try:
                    st.session_state["current_case_id"] = _save_own_case("Assessed Together")
                    st.session_state["last_assessment_type"] = "Assessed Together"
                    go_to("dashboard")
                except ConnectionError as exc:
                    show_gs_error(exc)

    else:  # _mode == "self"
        if st.button("🔗 Generate Pre-Filled Magic Link for Attending", type="primary", width="stretch"):
            if not _assessment_has_value():
                st.warning("Please provide at least one rating or comment before generating a link.")
            else:
                try:
                    st.session_state["current_case_id"] = _save_own_case("Self-Assessment")
                    st.session_state["last_assessment_type"] = "Self-Assessment"
                    draft_id = save_draft(
                        resident_email=st.session_state["resident"],
                        date=st.session_state["date"],
                        specialty_id=st.session_state["specialty_id"],
                        procedure_id=st.session_state["procedure_id"],
                        attending_id=st.session_state["attending_id"],
                        scores_dict=st.session_state["scores"],
                        case_complexity=st.session_state["case_complexity"],
                        case_preparation=st.session_state["case_preparation"],
                        overall_performance=st.session_state["overall_performance"],
                        robo_type=_robo_type_to_save,
                        notes=st.session_state.get("notes", ""),
                        improve=st.session_state.get("improve", ""),
                        how=st.session_state.get("how", ""),
                    )
                    _att_match = atnd_df[atnd_df["attending_id"].astype(str).str.strip()
                                          == str(st.session_state.get("attending_id", "")).strip()]
                    safe_att = _att_match["attending_name"].values[0].replace(" ", "_") if len(_att_match) > 0 else "Unknown"
                    base_url = st.secrets.get("APP_BASE_URL", "https://procedurepassport.streamlit.app")
                    st.session_state["generated_magic_link"] = (
                        f"{base_url}/?mode=attending"
                        f"&resident={st.session_state['resident']}"
                        f"&procedure_id={st.session_state['procedure_id']}"
                        f"&specialty_id={st.session_state['specialty_id']}"
                        f"&attending_name={safe_att}"
                        f"&draft_id={draft_id}"
                    )
                    go_to("magic_link_ready")
                except ConnectionError as exc:
                    show_gs_error(exc)


# ════════════════════════════════════════════════════════════
# PAGE: SINGLE-CASE DASHBOARD
# ════════════════════════════════════════════════════════════
elif page == "dashboard":
    try:
        _, proc_df, steps_df, _ = load_refs()
    except ConnectionError as exc:
        show_gs_error(exc)
        st.stop()

    steps = steps_df[steps_df["procedure_id"] == st.session_state["procedure_id"]].sort_values("step_order")
    _dash_proc_rows = proc_df.loc[proc_df["procedure_id"] == st.session_state["procedure_id"], "procedure_name"].values
    _dash_proc_name = _dash_proc_rows[0] if len(_dash_proc_rows) else ""

    page_header("✅ Case Saved")
    st.success(f"Case ID: `{st.session_state.get('current_case_id', '—')}`")

    data = [{"Step": row["step_name"],
             "Rating": st.session_state["scores"].get(row["step_id"], "")}
            for _, row in steps.iterrows()]
    df   = pd.DataFrame(data)
    st.dataframe(style_df(df, "Rating"), width="stretch")

    meta_col1, meta_col2 = st.columns(2)
    with meta_col1:
        st.markdown(f"**Date:** {fmt_date(st.session_state.get('date', ''))}")
        st.markdown(f"**Case Complexity:** {st.session_state.get('case_complexity', '—')}")
        st.markdown(f"**Daily Preparation:** {st.session_state.get('case_preparation', '—')}")
    with meta_col2:
        st.markdown(f"**Overall Performance:** {st.session_state.get('overall_performance', '—')}")
        st.markdown(f"**Basis:** {st.session_state.get('last_assessment_type', '—')}")
        if _is_robotic_procedure(_dash_proc_name):
            st.markdown(f"**Robot:** {st.session_state.get('robo_type', '—')}")

    if st.session_state.get("improve", "").strip() or st.session_state.get("how", "").strip():
        st.markdown(f"**In order to improve this:** {st.session_state.get('improve', '') or '_(blank)_'}.")
        st.markdown(f"**Do this:** {st.session_state.get('how', '') or '_(blank)_'}.")

    if st.session_state.get("notes", "").strip():
        st.markdown("**Comments:**")
        st.info(st.session_state["notes"])

    st.markdown("---")
    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("⬅️ Back to Assessment"):
            go_to("assessment")
    with col2:
        if st.button("🏠 Home"):
            go_to("home")
    with col3:
        if st.button("➕ New Assessment", type="primary"):
            go_to("start")


# ════════════════════════════════════════════════════════════
# PAGE: MAGIC LINK READY (after a self-assessment generates one)
# ════════════════════════════════════════════════════════════
elif page == "magic_link_ready":
    if not st.session_state.get("generated_magic_link"):
        st.error("No magic link found. Please generate one from the assessment page.")
        if st.button("⬅️ Back to Assessment"):
            go_to("assessment")
        st.stop()

    page_header("🔗 Magic Link Ready")
    st.success("✅ Your self-assessment was saved, and a pre-filled link is ready for your attending:")
    copy_link_button(st.session_state["generated_magic_link"], key="copy_generated_link")
    st.code(st.session_state.get("generated_magic_link", ""), language="text")
    st.caption("The attending can review and adjust every field before submitting.")

    st.markdown("---")
    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("⬅️ Back to Assessment"):
            go_to("assessment")
    with col2:
        if st.button("🏠 Home"):
            go_to("home")
    with col3:
        if st.button("➕ New Assessment", type="primary"):
            go_to("start")


# ════════════════════════════════════════════════════════════
# PAGE: COMMENTS DASHBOARD
# ════════════════════════════════════════════════════════════
elif page == "comments":
    page_header("💬 Comments Dashboard")
    if st.button("🏠 Back to Home", key="comments_home_top"):
        go_to("home")
    resident = st.session_state.get("resident")
    if not resident:
        st.error("Not logged in.")
        if st.button("⬅️ Back to Home"):
            go_to("home")
        st.stop()

    try:
        merged = _build_resident_comments_df(resident)
    except ConnectionError as exc:
        show_gs_error(exc)
        if st.button("⬅️ Back to Home"):
            go_to("home")
        st.stop()

    if merged.empty:
        st.info("No comments recorded yet.")
        if st.button("⬅️ Back to Home"):
            go_to("home")
    else:
        st.caption("💡 Tip: To screenshot the full table — on mobile use print preview; on desktop use File > Print (Cmd+P / Ctrl+P), then adjust the scale percentage down until all columns fit on one page before screenshotting.")

        # Fix 8: procedure/attending filter dropdowns — each one's options
        # are narrowed by the OTHER dropdown's current selection, so e.g.
        # filtering to a procedure leaves only the attendings who have
        # entries for it in the Attending dropdown.
        _proc_selected = st.session_state.get("comments_proc_filter", "All Procedures")
        _att_selected = st.session_state.get("comments_att_filter", "All Attendings")

        _proc_pool = merged if _att_selected == "All Attendings" else merged[merged["Attending"] == _att_selected]
        _proc_opts = ["All Procedures"] + sorted(_proc_pool["Procedure"].dropna().unique().tolist())
        _att_pool = merged if _proc_selected == "All Procedures" else merged[merged["Procedure"] == _proc_selected]
        _att_opts = ["All Attendings"] + sorted(
            _att_pool["Attending"].dropna().unique().tolist(),
            key=lambda n: n.split()[-1] if n.split() else n,
        )

        # A previously-selected filter value can fall out of the newly
        # narrowed options (because the other filter now excludes it) —
        # reset it to "All ..." before the widget renders, rather than
        # letting st.selectbox raise on a default no longer in its options.
        if _proc_selected not in _proc_opts:
            st.session_state["comments_proc_filter"] = "All Procedures"
        if _att_selected not in _att_opts:
            st.session_state["comments_att_filter"] = "All Attendings"

        _filter_col1, _filter_col2 = st.columns(2)
        with _filter_col1:
            _proc_filter = st.selectbox("Filter by Procedure", _proc_opts, key="comments_proc_filter")
        with _filter_col2:
            _att_filter = st.selectbox("Filter by Attending", _att_opts, key="comments_att_filter")
        if _proc_filter != "All Procedures":
            merged = merged[merged["Procedure"] == _proc_filter]
        if _att_filter != "All Attendings":
            merged = merged[merged["Attending"] == _att_filter]

        # Filtering down to a single procedure/attending makes that column
        # redundant (every row shows the same value) — drop it from the
        # on-screen table while filtered.
        _show_proc = _proc_filter == "All Procedures"
        _show_att = _att_filter == "All Attendings"

        _render_comments_html_table(merged, _show_proc, _show_att)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            merged.drop(columns=["Comments_html"]).to_excel(writer, index=False, sheet_name="Comments")
        st.download_button(
            label="📥 Download as Excel",
            data=output.getvalue(),
            file_name=f"{resident}_comments.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        if st.button("⬅️ Back to Home"):
            go_to("home")


# ════════════════════════════════════════════════════════════
# PAGE: CUMULATIVE DASHBOARD
# ════════════════════════════════════════════════════════════
elif page == "cumulative":
    mobile_tip("📱 On mobile: tap the >> icon at top left to view the sidebar.")
    page_header("📊 Cumulative Dashboard")
    if st.button("🏠 Back to Home", key="cumulative_home_top"):
        go_to("home")
    resident = st.session_state.get("resident")
    if not resident:
        st.error("Not logged in.")
        if st.button("⬅️ Back to Home"):
            go_to("home")
        st.stop()

    try:
        merged, steps_df, procs_map = _build_resident_case_matrix(resident)
    except ConnectionError as exc:
        show_gs_error(exc)
        if st.button("⬅️ Back to Home"):
            go_to("home")
        st.stop()

    if merged.empty:
        st.info("No cases logged yet.")
        if st.button("⬅️ Back to Home"):
            go_to("home")
        st.stop()

    # ── Procedure selector ────────────────────────────────
    proc_ids      = merged["case_procedure_id"].dropna().unique()
    selected_proc = st.selectbox(
        "Procedure",
        options=sorted(proc_ids, key=lambda x: procs_map.get(x, x)),
        format_func=lambda x: procs_map.get(x, x),
    )

    _render_resident_heatmap(merged, steps_df, procs_map, selected_proc, filename_stub=resident)

    if st.button("⬅️ Back to Home"):
        go_to("home")


# ════════════════════════════════════════════════════════════
# PAGE: ATTENDING START ASSESSMENT
# ════════════════════════════════════════════════════════════
elif page == "attending_start":
    mobile_tip("📱 On mobile: tap the >> icon at top left to view the sidebar.")
    page_header("📋 Start Assessment")
    if st.button("🏠 Back to Home", key="att_start_home_top"):
        go_to("attending_home")

    specialty_id = st.session_state.get("attending_login_specialty_id")
    if not specialty_id:
        st.error("No specialty assigned. Contact an admin.")
        st.stop()

    try:
        _, proc_df, _, _ = load_refs()
        residents_df = read_sheet_df(
            SHEET_RESIDENTS, expected_cols=["email", "name", "specialty_id", "created_at"]
        )
    except ConnectionError as exc:
        show_gs_error(exc)
        if st.button("⬅️ Back to Home", key="att_start_home_err"):
            go_to("attending_home")
        st.stop()

    my_residents = residents_df[residents_df["specialty_id"] == specialty_id]
    procs = proc_df[proc_df["specialty_id"] == specialty_id]

    if my_residents.empty:
        st.warning("⚠️ No residents configured for your specialty.")
        if st.button("⬅️ Back to Home", key="att_start_no_res"):
            go_to("attending_home")
        st.stop()
    if procs.empty:
        st.warning("⚠️ No procedures configured for your specialty.")
        if st.button("⬅️ Back to Home", key="att_start_no_proc"):
            go_to("attending_home")
        st.stop()

    res_map  = dict(zip(my_residents["name"], my_residents["email"]))
    proc_map = dict(zip(procs["procedure_name"], procs["procedure_id"]))

    _CHOOSE_RES  = "Choose Resident"
    _CHOOSE_PROC = "Choose Procedure"

    resident_choice = st.selectbox(
        "Resident",
        [_CHOOSE_RES] + sorted(res_map.keys(), key=lambda n: n.split()[-1] if n.split() else n),
    )
    procedure_choice = st.selectbox("Procedure", [_CHOOSE_PROC] + _ordered_procedure_names(proc_map))
    case_date = st.date_input("Date", st.session_state["date"])
    st.session_state["date"] = case_date

    resident_chosen  = resident_choice != _CHOOSE_RES
    procedure_chosen = procedure_choice != _CHOOSE_PROC

    if not (resident_chosen and procedure_chosen):
        st.info("Choose a resident and a procedure to continue.")

    st.markdown("---")

    if st.button("Start Assessment", type="primary", width="stretch",
                 disabled=not (resident_chosen and procedure_chosen)):
        # Reuses the same session keys — and the same blank assessment
        # page — the anonymous "Blank Magic Link" flow feeds into, just
        # populated directly instead of via a link's query params.
        st.session_state["resident"]            = res_map[resident_choice]
        st.session_state["procedure_id"]        = proc_map[procedure_choice]
        st.session_state["specialty_id"]        = specialty_id
        st.session_state["attending_name"]      = st.session_state.get("attending_login_name", "").replace(" ", "_")
        st.session_state["draft_id"]            = ""
        st.session_state["attending_link_date"] = str(case_date)
        go_to("attending_assessment")

    st.markdown("---")
    if st.button("⬅️ Back to Home", key="att_start_bottom_home"):
        go_to("attending_home")


# ════════════════════════════════════════════════════════════
# PAGE: ATTENDING RESIDENT DASHBOARD
# ════════════════════════════════════════════════════════════
elif page == "attending_resident_dashboard":
    mobile_tip("📱 On mobile: tap the >> icon at top left to view the sidebar.")
    page_header("📊 Resident Dashboard")
    if st.button("🏠 Back to Home", key="att_dash_home_top"):
        go_to("attending_home")

    specialty_id = st.session_state.get("attending_login_specialty_id")
    if not specialty_id:
        st.error("No specialty assigned. Contact an admin.")
        st.stop()

    try:
        _, proc_df, _, _ = load_refs()
        residents_df = read_sheet_df(
            SHEET_RESIDENTS, expected_cols=["email", "name", "specialty_id", "created_at"]
        )
        cases_df = read_sheet_df(
            SHEET_CASES,
            expected_cols=["case_id", "resident_email", "specialty_id", "procedure_id", "assessment_type"],
        )
    except ConnectionError as exc:
        show_gs_error(exc)
        if st.button("⬅️ Back to Home", key="att_dash_home_err"):
            go_to("attending_home")
        st.stop()

    # Same "attending-confirmed" definition used everywhere else on this
    # page (comments table, case matrix): a resident's own unsubmitted
    # self-assessment doesn't count as data an attending can review yet.
    confirmed_cases = cases_df[
        cases_df["assessment_type"].fillna("").astype(str).str.strip() != "Self-Assessment"
    ]
    my_residents = residents_df[
        (residents_df["specialty_id"] == specialty_id)
        & (residents_df["email"].isin(set(confirmed_cases["resident_email"])))
    ]
    if my_residents.empty:
        st.warning("⚠️ No residents with recorded cases in your specialty yet.")
        if st.button("⬅️ Back to Home", key="att_dash_no_res"):
            go_to("attending_home")
        st.stop()

    res_map = dict(zip(my_residents["name"], my_residents["email"]))
    _CHOOSE_RES = "Choose Resident"
    _res_col, _proc_col = st.columns(2)
    with _res_col:
        resident_choice = st.selectbox(
            "Resident",
            [_CHOOSE_RES] + sorted(res_map.keys(), key=lambda n: n.split()[-1] if n.split() else n),
            key="att_dash_resident",
        )
    if resident_choice == _CHOOSE_RES:
        with _proc_col:
            st.selectbox("Procedure (optional)", ["Choose a resident first"], disabled=True)
        st.info("Choose a resident to view their comments and progress.")
        st.stop()

    resident_email = res_map[resident_choice]

    # Fetched once here and reused below at the heatmap section — same
    # data, no need to ask for it twice.
    try:
        case_matrix, steps_df, procs_map = _build_resident_case_matrix(resident_email)
    except ConnectionError as exc:
        show_gs_error(exc)
        st.stop()

    # Only list procedures with actual step-level ratings data — i.e. ones
    # that would actually produce a heatmap, not just an attending-
    # confirmed case whose steps were all left "Not Assessed" (see
    # _build_resident_case_matrix's own docstring). procedure_id is
    # compared as a string, not the raw dtype — the cases and procedures
    # sheets can disagree on int vs. float vs. string for the same ID
    # (see _norm_id above), so a raw .isin() can silently drop real
    # matches.
    resident_procedure_ids = set(case_matrix["case_procedure_id"]) if not case_matrix.empty else set()
    procs = proc_df[
        (proc_df["specialty_id"] == specialty_id)
        & (proc_df["procedure_id"].astype(str).isin(resident_procedure_ids))
    ]
    proc_map = dict(zip(procs["procedure_name"], procs["procedure_id"]))
    _ALL_PROCS = "Choose Procedure for Heat Map"

    if not proc_map:
        # No procedure for this resident has step-level ratings data — the
        # dropdown would otherwise offer nothing but its own placeholder.
        with _proc_col:
            st.selectbox("Procedure (optional)", ["No heat maps available"], disabled=True)
        st.info(f"📊 No heat maps available for {resident_choice} yet — no step-level ratings recorded.")
        procedure_choice, procedure_id = _ALL_PROCS, None
    else:
        _proc_opts = [_ALL_PROCS] + _ordered_procedure_names(proc_map)
        # A procedure picked for the previous resident can fall outside this
        # resident's options — reset it rather than letting st.selectbox
        # raise on a default no longer in its options (same guard as the
        # Comments Dashboard's Procedure/Attending filters above).
        if st.session_state.get("att_dash_procedure") not in _proc_opts:
            st.session_state["att_dash_procedure"] = _ALL_PROCS
        with _proc_col:
            procedure_choice = st.selectbox(
                "Procedure (optional)",
                _proc_opts,
                key="att_dash_procedure",
            )
        procedure_id = proc_map.get(procedure_choice) if procedure_choice != _ALL_PROCS else None

    # Whether the Comments table is narrowed to the chosen procedure or
    # showing every procedure, independent of the dropdown itself — the
    # dropdown also drives the heatmap below, so switching it back to "All
    # Procedures" just to see every comment would lose the heatmap too.
    # The toggle always starts filtered whenever the resident/procedure
    # selection changes, rather than carrying over a stale "show all" from
    # a previous procedure.
    _comments_scope = f"{resident_email}|{procedure_choice}"
    if st.session_state.get("att_dash_comments_scope") != _comments_scope:
        st.session_state["att_dash_comments_scope"] = _comments_scope
        st.session_state["att_dash_show_all_comments"] = False
    show_all_comments = st.session_state["att_dash_show_all_comments"] or not procedure_id

    st.caption("💡 Tip: this page is print-friendly — on desktop use File > Print (Cmd+P / Ctrl+P); on mobile use print preview.")
    st.markdown("---")

    try:
        all_comments_df = _build_resident_comments_df(resident_email)
    except ConnectionError as exc:
        show_gs_error(exc)
        st.stop()

    if all_comments_df.empty:
        # Nothing to filter or toggle — skip the section header/button
        # entirely rather than showing controls over an empty table.
        st.info("💬 No comments recorded for this resident yet.")
    else:
        st.markdown(f"### 💬 Comments — {resident_choice}")
        if procedure_id:
            _toggle_label = "Show All Comments" if not show_all_comments else f"Show Only {procedure_choice} Comments"
            if st.button(_toggle_label, key="att_dash_comments_toggle"):
                st.session_state["att_dash_show_all_comments"] = not show_all_comments
                st.rerun()

        comments_df = all_comments_df
        if procedure_id and not show_all_comments:
            comments_df = comments_df[comments_df["Procedure"] == procedure_choice]

        if comments_df.empty:
            st.info("No comments recorded yet.")
        else:
            _render_comments_html_table(comments_df, show_proc=show_all_comments, show_att=True)

    if procedure_id:
        st.markdown("---")
        # case_matrix was already fetched above, and the dropdown only
        # ever offers a procedure_id that's in it, so it's never empty
        # here.
        _render_resident_heatmap(
            case_matrix, steps_df, procs_map, procedure_id,
            filename_stub=resident_choice.replace(" ", "_"),
        )

    st.markdown("---")
    if st.button("⬅️ Back to Home", key="att_dash_bottom_home"):
        go_to("attending_home")


# ════════════════════════════════════════════════════════════
# PAGE: ATTENDING ASSESSMENT (magic link)
# ════════════════════════════════════════════════════════════
elif page == "attending_assessment":
    resident_email = st.session_state.get("resident", "")
    procedure_id   = st.session_state.get("procedure_id", "")
    specialty_id   = st.session_state.get("specialty_id", "")
    attending_name = st.session_state.get("attending_name", "Unknown")

    if not (resident_email and procedure_id and specialty_id):
        st.error("⚠️ Missing required information in this link. Please ask the resident to resend.")
        st.stop()

    # Decode URL-safe attending name
    display_attending = attending_name.replace("_", " ")

    # Pre-fill from the resident's self-assessment draft, if this link carries one.
    draft_id = st.session_state.get("draft_id", "")
    _draft   = load_draft(draft_id) if draft_id else None

    try:
        _, proc_df_att, steps_df, _ = load_refs()
    except ConnectionError as exc:
        show_gs_error(exc)
        st.stop()

    # Resolve procedure name for display (Fix 3)
    _att_proc_rows = proc_df_att.loc[proc_df_att["procedure_id"] == procedure_id, "procedure_name"].values
    _att_proc_name = _att_proc_rows[0] if len(_att_proc_rows) else procedure_id

    # Resolve the resident's display name for the header, same "...
    # Assessment for {name}" phrasing the resident's own Assess
    # Together/Self-Assess page uses, so this page reads as the same
    # page. The magic link only carries the resident's email, not their
    # name, so look it up; fall back to the email if that fails.
    try:
        _residents_df = read_sheet_df(
            SHEET_RESIDENTS, expected_cols=["email", "name", "specialty_id", "created_at"]
        )
        _resident_match = _residents_df.loc[
            _residents_df["email"].astype(str).str.strip().str.lower() == resident_email.strip().lower()
        ]
        _resident_display_name = (
            _resident_match["name"].values[0] if len(_resident_match) else resident_email
        )
    except ConnectionError:
        _resident_display_name = resident_email

    page_header(
        header_break_before(f"📝 {_att_proc_name}", f"Assessment for {_resident_display_name}"),
        tier_text=f"📝 {_att_proc_name} Assessment",
    )
    if _draft:
        # One deliberate break point (see header_break_before): the
        # script below tries the whole notice on one line first, only
        # falling back to two lines — starting with "Review and
        # adjust..." — once shrinking the font hits a floor.
        _notice_text = header_break_before(
            "📋 This form has been pre-filled from the resident's self-assessment.",
            "Review and adjust before submitting.",
        )
        st.markdown(
            '<div class="pp-prefill-notice-wrap" style="border: 2px solid #1E88E5; '
            'background-color: #FFFFFF; color: #000000; border-radius: 0.5rem; '
            'padding: 0.75rem 1rem; margin-bottom: 0.75rem;">'
            f'<span class="pp-prefill-notice-text">{html.escape(_notice_text)}</span>'
            "</div>",
            unsafe_allow_html=True,
        )
        st.iframe(
            """
            <script>
            (function() {
                var doc = window.parent.document;
                var wraps = doc.querySelectorAll('.pp-prefill-notice-wrap');
                var wrap = wraps[wraps.length - 1];
                if (!wrap) return;
                var el = wrap.querySelector('.pp-prefill-notice-text');
                if (!el) return;
                var oneLineMaxPx = 16;   // 1rem ceiling on one line
                var floorPx = 12;        // 0.75rem — below this, switch to two lines instead
                var twoLineMaxPx = 15;   // 0.9375rem ceiling once wrapped
                // Measures a string's rendered single-line width at a
                // given font size via a detached, invisible probe.
                function measureWidth(str, fontPx) {
                    var probe = doc.createElement('span');
                    probe.style.position = 'absolute';
                    probe.style.visibility = 'hidden';
                    probe.style.whiteSpace = 'nowrap';
                    probe.style.fontSize = fontPx + 'px';
                    var computed = window.parent.getComputedStyle(el);
                    probe.style.fontFamily = computed.fontFamily;
                    probe.style.fontWeight = computed.fontWeight;
                    probe.textContent = str;
                    doc.body.appendChild(probe);
                    var w = probe.scrollWidth;
                    doc.body.removeChild(probe);
                    return w;
                }
                function fit() {
                    var containerWidth = wrap.clientWidth;
                    if (!containerWidth) return;
                    var fullText = el.textContent;
                    // header_break_before() leaves exactly one regular
                    // space (everything else is nbsp) — that's the one
                    // deliberate wrap point, right before "Review...".
                    var breakIdx = fullText.indexOf(' ');
                    var oneLineWidth = measureWidth(fullText, oneLineMaxPx);
                    if (oneLineWidth <= containerWidth) {
                        el.style.whiteSpace = 'nowrap';
                        el.style.fontSize = oneLineMaxPx + 'px';
                        return;
                    }
                    var oneLineFit = Math.max(1, oneLineMaxPx * (containerWidth / oneLineWidth));
                    if (oneLineFit >= floorPx) {
                        el.style.whiteSpace = 'nowrap';
                        el.style.fontSize = (oneLineFit * 0.96) + 'px';
                        return;
                    }
                    // Doesn't comfortably fit one line even shrunk to the
                    // floor — allow the wrap, sized so each of the two
                    // resulting lines fits its own width.
                    el.style.whiteSpace = 'normal';
                    var widest = breakIdx > -1
                        ? Math.max(
                            measureWidth(fullText.slice(0, breakIdx), twoLineMaxPx),
                            measureWidth(fullText.slice(breakIdx + 1), twoLineMaxPx)
                          )
                        : measureWidth(fullText, twoLineMaxPx);
                    var finalPx = widest <= containerWidth
                        ? twoLineMaxPx
                        : Math.max(1, twoLineMaxPx * (containerWidth / widest) * 0.9);
                    el.style.fontSize = finalPx + 'px';
                }
                fit();
                window.parent.addEventListener('resize', fit);
                if (window.parent.ResizeObserver) {
                    new window.parent.ResizeObserver(fit).observe(wrap);
                }
            })();
            </script>
            """,
            height=1,
        )
    assessment_instructions_note()

    if _is_robotic_procedure(_att_proc_name):
        robo_type = render_robo_type_picker("robo_type", default=(_draft or {}).get("robo_type", "Xi"))
    else:
        robo_type = None

    st.markdown("---")

    steps = steps_df[steps_df["procedure_id"] == procedure_id].sort_values("step_order")
    if steps.empty:
        st.error("This procedure has no defined steps. Please contact the program coordinator.")
        st.stop()

    # Defaults sourced from the draft when this link was pre-filled, else
    # the usual blanks — same fallback pattern the resident's own page uses.
    _d = _draft or {}
    # Resident:/Procedure:/Attending:/Date of Procedure used to be shown
    # here — dropped so this page matches the resident's own assessment
    # page, which doesn't show them either. The date is no longer
    # editable by the attending: it comes from the draft (self-assess
    # flow) or, for a blank link, the date the resident actually chose on
    # the Start page before generating it (attending_link_date, carried
    # by the link itself) — falling back to today only if neither is
    # present (e.g. an old link from before this was added).
    _date_source = _d.get("date") or st.session_state.get("attending_link_date")
    case_date = datetime.date.today()
    if _date_source:
        try:
            case_date = datetime.date.fromisoformat(str(_date_source)[:10])
        except ValueError:
            pass

    with st.container(key="assess_improve_how"):
        _att_imp_label_col, _att_imp_input_col = st.columns([2, 6])
        with _att_imp_label_col:
            st.markdown(
                '<p style="text-align: left;">In order to improve this:</p>',
                unsafe_allow_html=True,
            )
        with _att_imp_input_col:
            improve = st.text_input(
                "What to improve",
                value=_d.get("improve", ""),
                key="assess_improve",
                label_visibility="collapsed",
                placeholder="e.g., suture technique",
            )
        _att_how_label_col, _att_how_input_col = st.columns([2, 6])
        with _att_how_label_col:
            st.markdown("Do this:")
        with _att_how_input_col:
            how = st.text_input(
                "How to improve it",
                value=_d.get("how", ""),
                key="assess_how",
                label_visibility="collapsed",
                placeholder="e.g., practice two-handed knots",
            )
    sync_improve_how_label_width()

    st.markdown("---")

    with st.container(key="assess_ratings_row"):
        _att_overall_col, _att_prep_col = st.columns(2)
        with _att_overall_col:
            _att_o_default = _d.get("overall_performance", O_SCORE_OPTIONS[0])
            _att_o_idx = O_SCORE_OPTIONS.index(_att_o_default) if _att_o_default in O_SCORE_OPTIONS else 0
            o_score = st.selectbox("Overall Performance Rating", O_SCORE_OPTIONS, index=_att_o_idx, key="assess_overall_performance")
        with _att_prep_col:
            _att_cp_opts = ["Not Assessed", "Unprepared", "Poorly Prepared",
                            "Adequately Prepared", "Well Prepared", "Highly Prepared"]
            _att_cp_default = _d.get("case_preparation", "Not Assessed")
            _att_cp_idx = _att_cp_opts.index(_att_cp_default) if _att_cp_default in _att_cp_opts else 0
            case_preparation = st.selectbox("Daily Preparation", _att_cp_opts, index=_att_cp_idx, key="assess_preparation")

    _att_cc_opts = ["— Select complexity —", "Straight Forward", "Moderate", "Complex"]
    _att_cc_default = _d.get("case_complexity", "— Select complexity —")
    _att_cc_idx = _att_cc_opts.index(_att_cc_default) if _att_cc_default in _att_cc_opts else 0

    # The resolved (index-0-fallback-aware) starting value for each field —
    # used below to check whether "Changes As Made Above" is actually true,
    # not just checked. Comparing against these instead of the raw draft
    # dict avoids false "changed" positives from e.g. a blank/NaN draft
    # value resolving to the same displayed default the widget already
    # falls back to on its own.
    _draft_resolved_o           = O_SCORE_OPTIONS[_att_o_idx]
    _draft_resolved_preparation = _att_cp_opts[_att_cp_idx]
    _draft_resolved_complexity  = _att_cc_opts[_att_cc_idx]

    scores: dict = {}
    _draft_scores = _d.get("scores") or {}
    _draft_resolved_scores: dict = {}
    with st.expander(
        header_break_before("Step-Level Ratings for", _att_proc_name),
        expanded=False,
        key="step_ratings_expander_attending",
    ):
        # Case Complexity leads the Step-Level Ratings section, then each
        # procedure step in order.
        case_complexity = st.selectbox(
            "Case Complexity", _att_cc_opts, index=_att_cc_idx, key="assess_case_complexity"
        )
        for _, row in steps.iterrows():
            step_id   = row["step_id"]
            step_name = row["step_name"]
            _step_default = _draft_scores.get(step_id, "Not Assessed")
            _step_idx = RATING_OPTIONS.index(_step_default) if _step_default in RATING_OPTIONS else 0
            _draft_resolved_scores[step_id] = RATING_OPTIONS[_step_idx]
            scores[step_id] = st.selectbox(
                step_name, RATING_OPTIONS, index=_step_idx, key=f"att_score_{step_id}"
            )

    st.markdown("---")

    notes = st.text_area(
        "Development / Improvement / Feed-Forward (optional)",
        value=_d.get("notes", ""),
        key="assess_notes",
    )

    _accept_no_changes   = False
    _accept_with_changes = False
    if _draft:
        st.markdown("---")
        st.markdown("**This form was pre-filled from the resident's self-assessment, please confirm:**")
        _accept_no_changes = st.checkbox(
            "No changes. Accept Resident Self-Assessment", key="assess_accept_no_changes"
        )
        _accept_with_changes = st.checkbox(
            "Changes As Made Above", key="assess_accept_with_changes"
        )

    st.markdown("---")
    if st.button("✅ Submit Evaluation", type="primary", width="stretch"):
        _has_value = (
            case_complexity != "— Select complexity —"
            or case_preparation != "Not Assessed"
            or o_score != O_SCORE_OPTIONS[0]
            or any(v != "Not Assessed" for v in scores.values())
            or notes.strip() != ""
            or improve.strip() != ""
            or how.strip() != ""
        )
        _matches_draft = (
            case_complexity == _draft_resolved_complexity
            and case_preparation == _draft_resolved_preparation
            and o_score == _draft_resolved_o
            and notes == _d.get("notes", "")
            and improve == _d.get("improve", "")
            and how == _d.get("how", "")
            and all(scores.get(sid) == _draft_resolved_scores.get(sid) for sid in scores)
        )
        if not _has_value:
            st.warning("Please provide at least one rating or comment before submitting.")
        elif _draft and not (_accept_no_changes or _accept_with_changes):
            st.warning("Please check one of the two boxes above before submitting.")
        elif _draft and _accept_no_changes and _accept_with_changes:
            st.warning("Please check only one of the two boxes above, not both.")
        elif _draft and _accept_with_changes and _matches_draft:
            st.warning("You checked “Changes As Made Above,” but nothing was actually "
                       "changed from the resident's self-assessment. Please make a change, "
                       "or check “No changes. Accept Resident Self-Assessment” instead.")
        else:
            if _draft:
                _assessment_type = ("Attending Evaluation (Accepted Self-Assessment, No Changes)"
                                     if _accept_no_changes else
                                     "Attending Evaluation (Pre-filled, Changes Made)")
            else:
                _assessment_type = "Attending Evaluation (Blank)"
            # A logged-in attending has a real attending_id — use it so the
            # case is attributed properly (filters/exports by Attending, a
            # future "my submissions" view, etc.). The anonymous magic-link
            # flow has no such account, so it keeps the decodable magic_
            # prefix instead.
            _attending_id_for_save = (
                st.session_state.get("attending_login_id")
                if st.session_state.get("role") == "attending" and st.session_state.get("attending_login_id")
                else f"magic_{attending_name}"
            )
            try:
                case_id = save_case(
                    resident_email=resident_email,
                    date=case_date,
                    specialty_id=specialty_id,
                    procedure_id=procedure_id,
                    attending_id=_attending_id_for_save,
                    scores_dict=scores,
                    notes=notes,
                    case_complexity=case_complexity,
                    case_preparation=case_preparation,
                    overall_performance=o_score,
                    robo_type=robo_type,
                    improve=improve,
                    how=how,
                    assessment_type=_assessment_type,
                )
                if draft_id:
                    delete_draft(draft_id)
                # Store submission summary for the confirmation page
                st.session_state["attending_submission"] = {
                    "case_id":             case_id,
                    "resident_email":      resident_email,
                    "resident_name":       _resident_display_name,
                    "procedure_id":        procedure_id,
                    "procedure_name":      _att_proc_name,
                    "attending_name":      display_attending,
                    "date":                str(case_date),
                    "case_complexity":     case_complexity,
                    "case_preparation":    case_preparation,
                    "overall_performance": o_score,
                    "robo_type":           robo_type,
                    "notes":               notes,
                    "improve":             improve,
                    "how":                 how,
                    "assessment_type":     _assessment_type,
                    "scores":              scores,
                    "steps":               steps[["step_id", "step_name"]].to_dict("records"),
                }
                go_to("attending_confirmation")
            except ConnectionError as exc:
                show_gs_error(exc)


# ════════════════════════════════════════════════════════════
# PAGE: ATTENDING CONFIRMATION
# ════════════════════════════════════════════════════════════
elif page == "attending_confirmation":
    sub = st.session_state.get("attending_submission")
    if not sub:
        st.error("No submission data found.")
        st.stop()

    page_header("✅ Evaluation Submitted")
    st.success("Thank you! Your evaluation has been recorded.")

    _robo_type_line = f'<br><b>Robot:</b> {sub["robo_type"]}' if sub.get("robo_type") else ""
    st.markdown(
        f'<div class="pp-card">'
        f'<b>Resident:</b> {sub.get("resident_name", sub["resident_email"])}<br>'
        f'<b>Attending:</b> {sub["attending_name"]}<br>'
        f'<b>Procedure:</b> {sub.get("procedure_name", sub["procedure_id"])}<br>'
        f'<b>Date:</b> {fmt_date(sub["date"])}<br>'
        f'<b>Overall Performance:</b> {sub["overall_performance"]}'
        f'{_robo_type_line}'
        f'</div>',
        unsafe_allow_html=True,
    )

    if sub.get("improve", "").strip() or sub.get("how", "").strip():
        st.markdown(f"**In order to improve this:** {sub.get('improve', '') or '_(blank)_'}.")
        st.markdown(f"**Do this:** {sub.get('how', '') or '_(blank)_'}.")

    if sub["notes"].strip():
        st.markdown("**Comments submitted:**")
        st.info(sub["notes"])

    st.markdown("#### Step Ratings Submitted")
    step_rows = []
    for step_rec in sub["steps"]:
        step_id   = step_rec["step_id"]
        step_name = step_rec["step_name"]
        rating    = sub["scores"].get(step_id, "—")
        step_rows.append({"Step": step_name, "Rating": rating})

    summary_df = pd.DataFrame(step_rows)
    st.dataframe(style_df(summary_df, "Rating"), width="stretch")

    st.markdown("---")
    if st.session_state.get("role") == "attending" and st.session_state.get("attending_login_email"):
        st.markdown("_The resident can view this evaluation in their dashboard._")
        _att_confirm_cols = st.columns(2)
        with _att_confirm_cols[0]:
            if st.button("➕ Start Another Assessment", type="primary", width="stretch", key="att_confirm_another"):
                go_to("attending_start")
        with _att_confirm_cols[1]:
            if st.button("🏠 Back to Home", width="stretch", key="att_confirm_home"):
                go_to("attending_home")
    else:
        st.markdown("_You may now close this window. The resident can view the evaluation in their dashboard._")


# Runs after every page render, regardless of which page/branch above
# executed, so it always fits whatever buttons ended up on screen.
fit_all_button_labels()
suppress_picker_keyboards()
